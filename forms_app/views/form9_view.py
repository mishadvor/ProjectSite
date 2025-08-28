# forms_app/views/form9_view.py
import os
import pandas as pd
import numpy as np
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from io import BytesIO
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


def form9_view(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("file")

        # Проверка файла
        if not uploaded_file:
            messages.error(request, "Файл не загружен.")
            return render(request, "forms_app/form9.html")

        if not uploaded_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Поддерживаются только файлы .xlsx")
            return render(request, "forms_app/form9.html")

        try:
            # === ЧИТАЕМ ФАЙЛ В ПАМЯТИ ===
            file_data = uploaded_file.read()
            df_raw = pd.read_excel(BytesIO(file_data), header=1)
            df_raw = df_raw.reset_index(drop=True)

            # Убедимся, что числовые колонки корректны
            df_raw["шт."] = pd.to_numeric(df_raw["шт."], errors="coerce")
            df_raw["Текущий остаток, шт."] = pd.to_numeric(
                df_raw["Текущий остаток, шт."], errors="coerce"
            )
            df_raw["Выкупили, шт."] = pd.to_numeric(
                df_raw["Выкупили, шт."], errors="coerce"
            )

            # Проверка колонки 'Склад'
            if "Склад" not in df_raw.columns:
                raise ValueError(
                    f"Колонка 'Склад' не найдена. Доступные: {df_raw.columns.tolist()}"
                )

            # Проверка обязательных колонок
            required_cols = ["Артикул WB", "Баркод", "Артикул продавца", "Размер"]
            missing_cols = [col for col in required_cols if col not in df_raw.columns]
            if missing_cols:
                messages.error(
                    request, f"Не хватает колонок: {', '.join(missing_cols)}"
                )
                return render(request, "forms_app/form9.html")

            # --- Лист 1: Оборот (без складов) ---
            df1 = (
                df_raw.groupby(
                    ["Артикул WB", "Баркод", "Артикул продавца", "Размер"],
                    as_index=False,
                )
                .agg(
                    {
                        "шт.": "sum",
                        "Текущий остаток, шт.": "sum",
                        "Выкупили, шт.": "sum",
                    }
                )
                .round(0)
            )
            df1 = df1.rename(columns={"шт.": "Заказы, шт."})

            # Оборачиваемость по заказам
            numerator = df1["Текущий остаток, шт."]
            denominator = df1["Заказы, шт."]
            conditions = [
                (numerator == 0) & (denominator == 0),
                (numerator == 0) & (denominator > 0),
                (numerator > 0) & (denominator == 0),
                (numerator > 0) & (denominator > 0),
            ]
            turnover_value = (
                (numerator / denominator * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value.astype(str),
            ]
            df1["Оборачиваемость по Заказам"] = np.select(
                conditions, choices, default="0"
            )

            # Оборачиваемость по Продажам
            numerator_sell = df1["Текущий остаток, шт."]
            denominator_sell = df1["Выкупили, шт."]
            conditions_sell = [
                (numerator_sell == 0) & (denominator_sell == 0),
                (numerator_sell == 0) & (denominator_sell > 0),
                (numerator_sell > 0) & (denominator_sell == 0),
                (numerator_sell > 0) & (denominator_sell > 0),
            ]
            turnover_value_sell = (
                (numerator_sell / denominator_sell * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices_sell = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value_sell.astype(str),
            ]
            df1["Оборачиваемость по Продажам"] = np.select(
                conditions_sell, choices_sell, default="0"
            )

            # Сортировка
            df1_orders = df1.sort_values(by=["Текущий остаток, шт."], ascending=False)
            df1_sales = df1.sort_values(by=["Текущий остаток, шт."], ascending=False)

            # --- Лист 2: Оборот по складам ---
            df2_grouped = (
                df_raw.groupby(
                    ["Артикул WB", "Баркод", "Артикул продавца", "Размер", "Склад"],
                    as_index=False,
                )
                .agg(
                    {
                        "шт.": "sum",
                        "Текущий остаток, шт.": "sum",
                        "Выкупили, шт.": "sum",
                    }
                )
                .round(0)
            )
            df2_grouped["Тип склада"] = df2_grouped["Склад"].apply(
                lambda x: (
                    "Мой склад" if x == "Склад поставщика - везу на склад WB" else "FBO"
                )
            )
            df2_grouped = df2_grouped.rename(columns={"шт.": "Заказы, шт."})

            # --- Лист 3: Оборот по артикулам (без размеров) ---
            df3 = (
                df_raw.groupby(
                    ["Артикул WB", "Артикул продавца"],
                    as_index=False,
                )
                .agg(
                    {
                        "шт.": "sum",
                        "Текущий остаток, шт.": "sum",
                        "Выкупили, шт.": "sum",
                    }
                )
                .round(0)
            )
            df3 = df3.rename(columns={"шт.": "Заказы, шт."})

            # Оборачиваемость по Заказам
            numerator3 = df3["Текущий остаток, шт."]
            denominator3 = df3["Заказы, шт."]
            conditions3 = [
                (numerator3 == 0) & (denominator3 == 0),
                (numerator3 == 0) & (denominator3 > 0),
                (numerator3 > 0) & (denominator3 == 0),
                (numerator3 > 0) & (denominator3 > 0),
            ]
            turnover_value3 = (
                (numerator3 / denominator3 * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices3 = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value3.astype(str),
            ]
            df3["Оборачиваемость по Заказам"] = np.select(
                conditions3, choices3, default="0"
            )

            # Оборачиваемость по Продажам
            numerator3_sell = df3["Текущий остаток, шт."]
            denominator3_sell = df3["Выкупили, шт."]
            conditions3_sell = [
                (numerator3_sell == 0) & (denominator3_sell == 0),
                (numerator3_sell == 0) & (denominator3_sell > 0),
                (numerator3_sell > 0) & (denominator3_sell == 0),
                (numerator3_sell > 0) & (denominator3_sell > 0),
            ]
            turnover_value3_sell = (
                (numerator3_sell / denominator3_sell * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices3_sell = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value3_sell.astype(str),
            ]
            df3["Оборачиваемость по Продажам"] = np.select(
                conditions3_sell, choices3_sell, default="0"
            )

            # Сортировка
            df3_orders = df3.sort_values(by=["Текущий остаток, шт."], ascending=False)
            df3_sales = df3.sort_values(by=["Текущий остаток, шт."], ascending=False)

            # Оборачиваемость по Заказам
            numerator2 = df2_grouped["Текущий остаток, шт."]
            denominator2 = df2_grouped["Заказы, шт."]
            conditions2 = [
                (numerator2 == 0) & (denominator2 == 0),
                (numerator2 == 0) & (denominator2 > 0),
                (numerator2 > 0) & (denominator2 == 0),
                (numerator2 > 0) & (denominator2 > 0),
            ]
            turnover_value2 = (
                (numerator2 / denominator2 * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices2 = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value2.astype(str),
            ]
            df2_grouped["Оборачиваемость по Заказам"] = np.select(
                conditions2, choices2, default="0"
            )

            # Оборачиваемость по Продажам
            numerator2_sell = df2_grouped["Текущий остаток, шт."]
            denominator2_sell = df2_grouped["Выкупили, шт."]
            conditions2_sell = [
                (numerator2_sell == 0) & (denominator2_sell == 0),
                (numerator2_sell == 0) & (denominator2_sell > 0),
                (numerator2_sell > 0) & (denominator2_sell == 0),
                (numerator2_sell > 0) & (denominator2_sell > 0),
            ]
            turnover_value2_sell = (
                (numerator2_sell / denominator2_sell * 7)
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )
            choices2_sell = [
                "0",
                "пополнить/FBS",
                "SOS!SOS!SOS!SOS!",
                turnover_value2_sell.astype(str),
            ]
            df2_grouped["Оборачиваемость по Продажам"] = np.select(
                conditions2_sell, choices2_sell, default="0"
            )

            df2_orders = df2_grouped.sort_values(
                by=["Текущий остаток, шт."], ascending=False
            )
            df2_sales = df2_grouped.sort_values(
                by=["Текущий остаток, шт."], ascending=False
            )

            # --- Функция: добавить градацию ---
            def add_turnover_grade(
                df, turnover_column, grade_column_prefix, is_sales=False
            ):
                """Добавляет колонку с градацией оборачиваемости"""
                df_copy = df.copy()
                df_copy["Оборачиваемость_num"] = pd.to_numeric(
                    df_copy[turnover_column], errors="coerce"
                )

                if is_sales:
                    # Градация для продаж
                    bins = [1, 39, 59, 79, 119, 179, 299, 499, float("inf")]
                else:
                    # Градация для заказов
                    bins = [0, 117, 177, 237, 357, 537, 897, 1497, float("inf")]

                labels = [
                    "Сильный дефицит",
                    "Средний дефицит",
                    "Дефицит 70%",
                    "Неактуальный товар 60%",
                    "Неактуальный товар 80%",
                    "Неактуальный товар 100%",
                    "Неликвид 80%",
                    "Неликвид 100%",
                ]

                df_copy[grade_column_prefix] = pd.cut(
                    df_copy["Оборачиваемость_num"],
                    bins=bins,
                    labels=labels,
                    right=False,
                    include_lowest=True,
                ).astype(str)

                df_copy[grade_column_prefix] = np.where(
                    df_copy[turnover_column] == "SOS!SOS!SOS!SOS!",
                    "SOS",
                    df_copy[grade_column_prefix],
                )

                df_copy = df_copy.drop(columns=["Оборачиваемость_num"], errors="ignore")
                return df_copy

            # --- Функция: форматирование с заливкой ---
            def format_sheet(sheet, grade_column_name="Градация по Заказам"):
                # Стиль заголовков
                style_name = "header_style"
                if style_name not in sheet.parent.named_styles:
                    header_style = NamedStyle(name=style_name)
                    header_style.font = Font(bold=True)
                    header_style.alignment = Alignment(
                        wrap_text=True, horizontal="center", vertical="center"
                    )
                    sheet.parent.add_named_style(header_style)

                for cell in sheet[1]:
                    cell.style = style_name

                # Проверяем, нужно ли пропустить цветовое выделение
                sheet_name = sheet.title
                skip_coloring_sheets = ["1. SOS по Заказам", "1. SOS по Продажам"]

                if sheet_name not in skip_coloring_sheets:
                    # Разные цветовые гаммы для разных типов градации
                    if "Продаж" in grade_column_name or "Продаж" in sheet_name:
                        colors = {
                            "SOS": "eb6a6a",
                            "Сильный дефицит": "f4f3a9",
                            "Средний дефицит": "d7e7bc",
                            "Дефицит 70%": "9cdaa6",
                            "Неактуальный товар 60%": "d7e7bc",
                            "Неактуальный товар 80%": "f4f3a9",
                            "Неактуальный товар 100%": "f4f3a9",
                            "Неликвид 80%": "ffa6a6",
                            "Неликвид 100%": "eb6a6a",
                        }
                    else:
                        colors = {
                            "SOS": "eb6a6a",
                            "Сильный дефицит": "f4f3a9",
                            "Средний дефицит": "d7e7bc",
                            "Дефицит 70%": "9cdaa6",
                            "Неактуальный товар 60%": "d7e7bc",
                            "Неактуальный товар 80%": "f4f3a9",
                            "Неактуальный товар 100%": "f4f3a9",
                            "Неликвид 80%": "ffa6a6",
                            "Неликвид 100%": "eb6a6a",
                        }

                    # Поиск колонки с градацией
                    header_row = [cell.value for cell in sheet[1]]
                    try:
                        grad_col_idx = header_row.index(grade_column_name) + 1
                    except ValueError:
                        grad_col_idx = None

                    # Раскраска строк
                    for row in sheet.iter_rows(min_row=2):
                        if grad_col_idx is None:
                            continue
                        grad_cell = row[grad_col_idx - 1]
                        value = str(grad_cell.value) if grad_cell.value else ""

                        fill_color = colors.get(value)
                        if fill_color:
                            fill = PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type="solid",
                            )
                            for cell in row:
                                if cell.value is not None:
                                    cell.fill = fill

                # Автоподбор ширины
                for column in sheet.columns:
                    max_length = 0
                    col_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value not in [None, ""]:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            continue
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[col_letter].width = adjusted_width

            # --- Шаг 2: Сохранение в один файл (в памяти) ---
            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl", mode="w") as writer:
                # ===== ОБОРОТ ОБЩИЙ =====
                df1_orders_final = add_turnover_grade(
                    df1_orders, "Оборачиваемость по Заказам", "Градация", is_sales=False
                )
                df1_orders_final = df1_orders_final[
                    [
                        "Артикул WB",
                        "Баркод",
                        "Артикул продавца",
                        "Размер",
                        "Заказы, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Заказам",
                        "Градация",
                    ]
                ]
                df1_orders_final.to_excel(
                    writer, index=False, sheet_name="Оборот_общий_Заказы"
                )

                df1_sales_final = add_turnover_grade(
                    df1_sales, "Оборачиваемость по Продажам", "Градация", is_sales=True
                )
                df1_sales_final = df1_sales_final[
                    [
                        "Артикул WB",
                        "Баркод",
                        "Артикул продавца",
                        "Размер",
                        "Выкупили, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Продажам",
                        "Градация",
                    ]
                ]
                df1_sales_final.to_excel(
                    writer, index=False, sheet_name="Оборот_общий_Продажи"
                )

                # ===== ОБОРОТ ПО СКЛАДАМ =====
                df2_orders_final = add_turnover_grade(
                    df2_orders, "Оборачиваемость по Заказам", "Градация", is_sales=False
                )
                df2_orders_final = df2_orders_final[
                    [
                        "Артикул WB",
                        "Баркод",
                        "Артикул продавца",
                        "Размер",
                        "Склад",
                        "Тип склада",
                        "Заказы, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Заказам",
                        "Градация",
                    ]
                ]
                df2_orders_final.to_excel(
                    writer, index=False, sheet_name="Оборот_по_складам_Заказы"
                )

                df2_sales_final = add_turnover_grade(
                    df2_sales, "Оборачиваемость по Продажам", "Градация", is_sales=True
                )
                df2_sales_final = df2_sales_final[
                    [
                        "Артикул WB",
                        "Баркод",
                        "Артикул продавца",
                        "Размер",
                        "Склад",
                        "Тип склада",
                        "Выкупили, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Продажам",
                        "Градация",
                    ]
                ]
                df2_sales_final.to_excel(
                    writer, index=False, sheet_name="Оборот_по_складам_Продажи"
                )

                # ===== ОБОРОТ ПО АРТИКУЛАМ (БЕЗ РАЗМЕРОВ) =====
                df3_orders_final = add_turnover_grade(
                    df3_orders, "Оборачиваемость по Заказам", "Градация", is_sales=False
                )
                df3_orders_final = df3_orders_final[
                    [
                        "Артикул WB",
                        "Артикул продавца",
                        "Заказы, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Заказам",
                        "Градация",
                    ]
                ]
                df3_orders_final.to_excel(
                    writer, index=False, sheet_name="Оборот_по_артикулам_Заказы"
                )

                df3_sales_final = add_turnover_grade(
                    df3_sales, "Оборачиваемость по Продажам", "Градация", is_sales=True
                )
                df3_sales_final = df3_sales_final[
                    [
                        "Артикул WB",
                        "Артикул продавца",
                        "Выкупили, шт.",
                        "Текущий остаток, шт.",
                        "Оборачиваемость по Продажам",
                        "Градация",
                    ]
                ]
                df3_sales_final.to_excel(
                    writer, index=False, sheet_name="Оборот_по_артикулам_Продажи"
                )

                # === Группировка по статусам для Оборачиваемости по Заказам ===
                df1_temp_orders = df1.copy()
                df1_temp_orders["Оборачиваемость_str"] = df1_temp_orders[
                    "Оборачиваемость по Заказам"
                ].astype(str)

                conditions_gr = [
                    df1_temp_orders["Оборачиваемость_str"] == "SOS!SOS!SOS!SOS!",
                    df1_temp_orders["Оборачиваемость_str"] == "пополнить/FBS",
                    df1_temp_orders["Оборачиваемость_str"] == "0",
                    pd.to_numeric(
                        df1_temp_orders["Оборачиваемость_str"], errors="coerce"
                    )
                    > 0,
                ]

                categories_oborot = [
                    "1. SOS по Заказам",
                    "2. пополнить_FBS по Заказам",
                    "3. 0 по Заказам",
                    "4. >0 по Заказам",
                ]

                df1_temp_orders["Группа по оборачиваемости"] = np.select(
                    conditions_gr, categories_oborot, default="Не попал"
                )

                # === Группировка по статусам для Оборачиваемости по Продажам ===
                df1_temp_sales = df1.copy()
                df1_temp_sales["Оборачиваемость_str"] = df1_temp_sales[
                    "Оборачиваемость по Продажам"
                ].astype(str)

                conditions_gr_sales = [
                    df1_temp_sales["Оборачиваемость_str"] == "SOS!SOS!SOS!SOS!",
                    df1_temp_sales["Оборачиваемость_str"] == "пополнить/FBS",
                    df1_temp_sales["Оборачиваемость_str"] == "0",
                    pd.to_numeric(
                        df1_temp_sales["Оборачиваемость_str"], errors="coerce"
                    )
                    > 0,
                ]

                categories_oborot_sales = [
                    "1. SOS по Продажам",
                    "2. пополнить_FBS по Продажам",
                    "3. 0 по Продажам",
                    "4. >0 по Продажам",
                ]

                df1_temp_sales["Группа по оборачиваемости"] = np.select(
                    conditions_gr_sales, categories_oborot_sales, default="Не попал"
                )

                # Запись групп для Заказов
                for category in categories_oborot:
                    filtered = df1_temp_orders[
                        df1_temp_orders["Группа по оборачиваемости"] == category
                    ]
                    if filtered.empty:
                        continue
                    filtered_with_grade = add_turnover_grade(
                        filtered,
                        "Оборачиваемость по Заказам",
                        "Градация",
                        is_sales=False,
                    )
                    filtered_clean = filtered_with_grade.drop(
                        columns=["Оборачиваемость_str", "Группа по оборачиваемости"],
                        errors="ignore",
                    )
                    cols_order = ["Артикул WB", "Баркод", "Артикул продавца", "Размер"]
                    other_cols = [
                        c for c in filtered_clean.columns if c not in cols_order
                    ]
                    filtered_clean = filtered_clean[cols_order + other_cols]
                    safe_sheet_name = category.replace("/", "_").replace("!", "")[:31]
                    filtered_clean.to_excel(
                        writer, sheet_name=safe_sheet_name, index=False
                    )

                # Запись групп для Продаж
                for category in categories_oborot_sales:
                    filtered = df1_temp_sales[
                        df1_temp_sales["Группа по оборачиваемости"] == category
                    ]
                    if filtered.empty:
                        continue
                    filtered_with_grade = add_turnover_grade(
                        filtered,
                        "Оборачиваемость по Продажам",
                        "Градация",
                        is_sales=True,
                    )
                    filtered_clean = filtered_with_grade.drop(
                        columns=["Оборачиваемость_str", "Группа по оборачиваемости"],
                        errors="ignore",
                    )
                    cols_order = ["Артикул WB", "Баркод", "Артикул продавца", "Размер"]
                    other_cols = [
                        c for c in filtered_clean.columns if c not in cols_order
                    ]
                    filtered_clean = filtered_clean[cols_order + other_cols]
                    safe_sheet_name = category.replace("/", "_").replace("!", "")[:31]
                    filtered_clean.to_excel(
                        writer, sheet_name=safe_sheet_name, index=False
                    )

                    # === Группировка по статусам для Оборачиваемости по Заказам (по артикулам) ===
                df3_temp_orders = df3.copy()
                df3_temp_orders["Оборачиваемость_str"] = df3_temp_orders[
                    "Оборачиваемость по Заказам"
                ].astype(str)

                conditions_gr3 = [
                    df3_temp_orders["Оборачиваемость_str"] == "SOS!SOS!SOS!SOS!",
                    df3_temp_orders["Оборачиваемость_str"] == "пополнить/FBS",
                    df3_temp_orders["Оборачиваемость_str"] == "0",
                    pd.to_numeric(
                        df3_temp_orders["Оборачиваемость_str"], errors="coerce"
                    )
                    > 0,
                ]

                categories_oborot3 = [
                    "1. SOS по Заказам (арт)",
                    "2. пополнить_FBS по Заказам (арт)",
                    "3. 0 по Заказам (арт)",
                    "4. >0 по Заказам (арт)",
                ]

                df3_temp_orders["Группа по оборачиваемости"] = np.select(
                    conditions_gr3, categories_oborot3, default="Не попал"
                )

                # === Группировка по статусам для Оборачиваемости по Продажам (по артикулам) ===
                df3_temp_sales = df3.copy()
                df3_temp_sales["Оборачиваемость_str"] = df3_temp_sales[
                    "Оборачиваемость по Продажам"
                ].astype(str)

                conditions_gr3_sales = [
                    df3_temp_sales["Оборачиваемость_str"] == "SOS!SOS!SOS!SOS!",
                    df3_temp_sales["Оборачиваемость_str"] == "пополнить/FBS",
                    df3_temp_sales["Оборачиваемость_str"] == "0",
                    pd.to_numeric(
                        df3_temp_sales["Оборачиваемость_str"], errors="coerce"
                    )
                    > 0,
                ]

                categories_oborot3_sales = [
                    "1. SOS по Продажам (арт)",
                    "2. пополнить_FBS по Продажам (арт)",
                    "3. 0 по Продажам (арт)",
                    "4. >0 по Продажам (арт)",
                ]

                df3_temp_sales["Группа по оборачиваемости"] = np.select(
                    conditions_gr3_sales, categories_oborot3_sales, default="Не попал"
                )

                # Запись групп для Заказов (по артикулам)
                for category in categories_oborot3:
                    filtered = df3_temp_orders[
                        df3_temp_orders["Группа по оборачиваемости"] == category
                    ]
                    if filtered.empty:
                        continue
                    filtered_with_grade = add_turnover_grade(
                        filtered,
                        "Оборачиваемость по Заказам",
                        "Градация",
                        is_sales=False,
                    )
                    filtered_clean = filtered_with_grade.drop(
                        columns=["Оборачиваемость_str", "Группа по оборачиваемости"],
                        errors="ignore",
                    )
                    cols_order = ["Артикул WB", "Артикул продавца"]
                    other_cols = [
                        c for c in filtered_clean.columns if c not in cols_order
                    ]
                    filtered_clean = filtered_clean[cols_order + other_cols]
                    safe_sheet_name = category.replace("/", "_").replace("!", "")[:31]
                    filtered_clean.to_excel(
                        writer, sheet_name=safe_sheet_name, index=False
                    )

                # Запись групп для Продаж (по артикулам)
                for category in categories_oborot3_sales:
                    filtered = df3_temp_sales[
                        df3_temp_sales["Группа по оборачиваемости"] == category
                    ]
                    if filtered.empty:
                        continue
                    filtered_with_grade = add_turnover_grade(
                        filtered,
                        "Оборачиваемость по Продажам",
                        "Градация",
                        is_sales=True,
                    )
                    filtered_clean = filtered_with_grade.drop(
                        columns=["Оборачиваемость_str", "Группа по оборачиваемости"],
                        errors="ignore",
                    )
                    cols_order = ["Артикул WB", "Артикул продавца"]
                    other_cols = [
                        c for c in filtered_clean.columns if c not in cols_order
                    ]
                    filtered_clean = filtered_clean[cols_order + other_cols]
                    safe_sheet_name = category.replace("/", "_").replace("!", "")[:31]
                    filtered_clean.to_excel(
                        writer, sheet_name=safe_sheet_name, index=False
                    )

                # Форматируем все листы
                workbook = writer.book
                for sheet_name in writer.sheets:
                    if "Заказ" in sheet_name:
                        format_sheet(workbook[sheet_name], "Градация")
                    elif "Продаж" in sheet_name:
                        format_sheet(workbook[sheet_name], "Градация")
                    else:
                        format_sheet(workbook[sheet_name], "Градация")

            # --- ОТПРАВКА ФАЙЛА КАК ОТВЕТ ---
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="oborot.xlsx"'
            return response

        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")
            return render(request, "forms_app/form9.html")

    else:
        from ..forms import ExcelProcessingForm

        form = ExcelProcessingForm()
        return render(request, "forms_app/form9.html", {"form": form})
