# forms_app/views/form11_view.py

import io
import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from django.http import FileResponse
import logging
import tempfile
import os

logger = logging.getLogger(__name__)

# --- Ваш mapping и функция process_sales_data остаются без изменений ---
склад_mapping = {
    # 1. Центральный федеральный округ (ЦФО)
    "Белая дача": "019_Белая дача",
    "Владимир": "013_Владимир",
    "Воронеж": "015_Воронеж",
    "Коледино": "012_Коледино",
    "Рязань (Тюшевское)": "014_Рязань (Тюшевское)",
    "Тула": "016_Тула",
    "Чашниково": "0191_Чашниково",
    "Электросталь": "011_Электросталь",
    "Котовск": "017_Котовск",
    "Склад поставщика - везу на склад WB": "018_Склад поставщика - везу на склад WB",
    "Подольск": "0190_Подольск",
    # 2. Северо-Западный федеральный округ (СЗФО)
    "Санкт-Петербург Уткина Заводь": "02_Санкт-Петербург Уткина Заводь",
    # 3. Южный федеральный округ (ЮФО)
    "Волгоград": "03_Волгоград",
    "Краснодар": "03_Краснодар",
    # 4. Северно-кавказкий федеральный округ (СКФО)
    "Невинномысск": "04_Невинномысск",
    # 5. Приволжский федеральный округ (ПФО)
    "Казань": "05_Казань",
    "Самара (Новосемейкино)": "05_Самара (Новосемейкино)",
    "Сарапул": "05_Сарапул",
    # 6. Уральский федеральный округ (УрФО)
    "Екатеринбург - Перспективный 12": "06_Екатеринбург - Перспективный 12",
    "Екатеринбург - Испытателей 14г": "06_Екатеринбург - Испытателей 14г",
    # 7. Сибирский федеральный округ (СФО)
    "Новосибирск": "07_Новосибирск",
    "СЦ Барнаул": "07_СЦ Барнаул",
    # 0. Города, не входящие в РФ:
    "Актобе": "0_Актобе",
    "Астана Карагандинское шоссе": "0_Астана",
    "Атакент": "0_Атакент",
    "СЦ Ереван": "0_СЦ Ереван",
}


def process_sales_data(df):
    """
    Обрабатывает DataFrame с данными о продажах согласно вашей логике.
    """
    # Удалить несколько колонок
    df = df.drop(
        columns=[
            "Бренд",
            "Предмет",
            "Сезон",
            "Коллекция",
            "Наименование",
            "Баркод",
            "Контракт",
            "Сумма заказов минус комиссия WB, руб.",
            "К перечислению за товар, руб.",
        ]
    )

    # --- Новый код для замены названий складов ---
    df["Склад"] = df["Склад"].map(склад_mapping).fillna(df["Склад"])

    # === Сортировка: сначала по 'Артикул продавца', затем внутри — по 'Склад', 'Размер' ===
    отсортированный_df_артикулы = df.sort_values(
        by=["Артикул продавца", "Склад", "Размер"], ascending=[True, True, True]
    )

    # --- Новый код для расчёта сумм по складам ---
    sum_zakazy_by_art_sku = отсортированный_df_артикулы.groupby(
        ["Артикул продавца", "Склад"]
    )["Заказы шт."].transform("sum")
    sum_vykup_by_art_sku = отсортированный_df_артикулы.groupby(
        ["Артикул продавца", "Склад"]
    )["Выкупили, шт."].transform("sum")
    sum_ostatok_by_art_sku = отсортированный_df_артикулы.groupby(
        ["Артикул продавца", "Склад"]
    )["Текущий остаток, шт."].transform("sum")

    отсортированный_df_артикулы["Сумма заказов, шт"] = sum_zakazy_by_art_sku
    отсортированный_df_артикулы["Сумма выкупили, шт"] = sum_vykup_by_art_sku
    отсортированный_df_артикулы["Сумма Текущий остаток, шт"] = sum_ostatok_by_art_sku

    # --- Новый код для расчёта "Наша оборачиваемость" ---
    отсортированный_df_артикулы["Наша оборачиваемость"] = (
        отсортированный_df_артикулы["Текущий остаток, шт."]
        .div(отсортированный_df_артикулы["Заказы шт."])
        .round(1)
        .where(
            отсортированный_df_артикулы["Заказы шт."].notna()
            & (отсортированный_df_артикулы["Заказы шт."] != 0)
        )
    )

    # --- Новый код для "Рекомендации для ФБО" ---
    import numpy as np

    отсортированный_df_артикулы["Рекомендации для ФБО"] = np.where(
        (отсортированный_df_артикулы["Наша оборачиваемость"] >= 0)
        & (отсортированный_df_артикулы["Наша оборачиваемость"] <= 2),
        "Рассмотреть",
        pd.NA,
    )

    # --- Новый код для вставки итогов и сортировки блоков ---
    grouped_with_sum = (
        отсортированный_df_артикулы.groupby("Артикул продавца", sort=False)
        .agg({"Заказы шт.": "sum"})
        .reset_index()
    )
    grouped_with_sum = grouped_with_sum.sort_values(by="Заказы шт.", ascending=False)
    sorted_artikuls = grouped_with_sum["Артикул продавца"].tolist()

    processed_parts = []
    for artikul in sorted_artikuls:
        current_group_df = отсортированный_df_артикулы[
            отсортированный_df_артикулы["Артикул продавца"] == artikul
        ]

        current_group_df = current_group_df.sort_values(
            by="Сумма заказов, шт", ascending=False, kind="stable"
        )

        processed_parts.append(current_group_df)

        total_zakazy = current_group_df["Заказы шт."].sum()
        total_vykup = current_group_df["Выкупили, шт."].sum()
        total_ostatok = current_group_df["Текущий остаток, шт."].sum()
        total_sum_zakazov = total_zakazy
        total_sum_vykup = total_vykup
        total_sum_ostatok = total_ostatok

        total_oborot_chisl = current_group_df["Текущий остаток, шт."].sum()
        total_oborot_znam = current_group_df["Заказы шт."].sum()
        avg_oborot = (
            (total_oborot_chisl / total_oborot_znam)
            if total_oborot_znam != 0
            else pd.NA
        )
        avg_oborot_rounded = round(avg_oborot, 1) if pd.notna(avg_oborot) else pd.NA

        recommendation_total = (
            "Рассмотреть"
            if (pd.notna(avg_oborot_rounded) and 0 <= avg_oborot_rounded <= 2)
            else pd.NA
        )

        total_row_data = {
            "Артикул WB": "Итого:",
            "Артикул продавца": artikul,
            "Размер": pd.NA,
            "Склад": pd.NA,
            "Заказы шт.": total_zakazy,
            "Выкупили, шт.": total_vykup,
            "Текущий остаток, шт.": total_ostatok,
            "Сумма заказов, шт": total_sum_zakazov,
            "Сумма выкупили, шт": total_sum_vykup,
            "Сумма Текущий остаток, шт": total_sum_ostatok,
            "Наша оборачиваемость": avg_oborot_rounded,
            "Рекомендации для ФБО": recommendation_total,
        }
        total_row_df = pd.DataFrame(
            [total_row_data], columns=отсортированный_df_артикулы.columns
        )
        processed_parts.append(total_row_df)

    # --- Изменение порядка колонок ---
    desired_order = [
        "Артикул WB",
        "Артикул продавца",
        "Размер",
        "Склад",
        "Заказы шт.",
        "Выкупили, шт.",
        "Текущий остаток, шт.",
        "Сумма заказов, шт",
        "Сумма выкупили, шт",
        "Сумма Текущий остаток, шт",
        "Наша оборачиваемость",
        "Рекомендации для ФБО",
    ]
    final_df = pd.concat(processed_parts, ignore_index=True)
    final_df = final_df[desired_order]

    return final_df


def apply_formatting(worksheet, processed_df):
    """
    Применяет форматирование к рабочему листу
    """
    # 1. Стиль заголовков
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=12)
    header_style.alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center"
    )
    header_style.fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )

    # Добавляем стиль в книгу, если его еще нет
    if "header_style" not in worksheet.parent.named_styles:
        worksheet.parent.add_named_style(header_style)

    # Применяем стиль к заголовкам
    for cell in worksheet[1]:
        cell.style = "header_style"

    # 2. Автоподбор ширины столбцов
    for column in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    # 3. Окрашивание строк по складам
    col_idx_sku = None
    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == "Склад":
            col_idx_sku = idx
            break

    if col_idx_sku:
        # Получаем уникальные значения "Склад" из processed_df
        unique_skus = processed_df["Склад"].dropna().unique()
        fill_map = {}
        colors = [
            "FFB6C1",  # Светло-розовый
            "90EE90",  # Светло-зеленый
            "87CEEB",  # Светло-голубой
            "FFFFE0",  # Светло-желтый
            "DDA0DD",  # Светло-фиолетовый
            "F0E68C",  # Хаки
            "FFA07A",  # Светло-лососевый
            "E6E6FA",  # Лавандовый
            "FFDAB9",  # Персиковый
            "20B2AA",  # Светло-морской волны
        ]

        for i, sku in enumerate(unique_skus):
            fill_map[sku] = PatternFill(
                start_color=colors[i % len(colors)],
                end_color=colors[i % len(colors)],
                fill_type="solid",
            )

        # Проходимся по строкам начиная с 2 (т.к. 1 - заголовки)
        for row_num in range(2, worksheet.max_row + 1):
            cell_in_sku_col = worksheet.cell(row=row_num, column=col_idx_sku)
            cell_value = cell_in_sku_col.value

            if cell_value in fill_map:
                fill_color = fill_map[cell_value]
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = fill_color

    # 4. Форматирование для строк "Итого:"
    for row_num in range(2, worksheet.max_row + 1):
        cell_in_art_wb = worksheet.cell(row=row_num, column=1)  # Колонка "Артикул WB"
        if cell_in_art_wb.value == "Итого:":
            # Жирный шрифт для строк "Итого:"
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = Font(bold=True)
                # Светло-серый фон для строк "Итого:"
                cell.fill = PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                )


def form11_view(request):
    """
    Представление для Формы 11.
    Обрабатывает загрузку Excel-файла, запускает обработку и возвращает результат.
    """
    if request.method == "POST" and request.FILES.get("excel_file"):
        uploaded_file = request.FILES["excel_file"]

        if not uploaded_file.name.lower().endswith(".xlsx"):
            context = {
                "error": "Пожалуйста, загрузите файл в формате .xlsx (поддерживаются .xlsx и .XLSX)."
            }
            return render(request, "forms_app/form11.html", context)

        try:
            df = pd.read_excel(uploaded_file, header=1)
            df = df.rename(columns={"шт.": "Заказы шт."})
            processed_df = process_sales_data(df)

            # Создаем временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                # Сохраняем данные в Excel с форматированием
                with pd.ExcelWriter(tmp_file.name, engine="openpyxl") as writer:
                    processed_df.to_excel(writer, sheet_name="Sheet1", index=False)

                    # Применяем форматирование
                    worksheet = writer.sheets["Sheet1"]
                    apply_formatting(worksheet, processed_df)

                # Используем FileResponse
                response = FileResponse(
                    open(tmp_file.name, "rb"),
                    as_attachment=True,
                    filename="рекомендации_фбо.xlsx",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # Удаляем временный файл после отправки
                def cleanup_temp_file():
                    try:
                        os.unlink(tmp_file.name)
                    except:
                        pass

                response.closed = cleanup_temp_file
                return response

        except Exception as e:
            logger.error(f"Ошибка при обработке файла в Форме 11: {e}")
            context = {"error": f"Произошла ошибка при обработке файла: {str(e)}"}
            return render(request, "forms_app/form11.html", context)

    return render(request, "forms_app/form11.html")
