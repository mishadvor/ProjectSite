# forms_app/views/form13_view.py

import pandas as pd
import tempfile
import os
from django.shortcuts import render
from django.http import FileResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def form13_simple_upload(request):
    """Форма 13: Признаки доступности - точное соответствие оригиналу"""
    if request.method == "POST" and request.FILES.get("file"):
        try:
            file = request.FILES["file"]

            # Загрузка данных (точно как в оригинале)
            df = pd.read_excel(
                file,
                sheet_name="Детальная информация",
                header=1,
                usecols=[
                    "Артикул продавца",
                    "Артикул WB",
                    "Размер",
                    "Доступность",
                    "Заказали, шт",
                    "Остатки на текущий день, шт",
                    "Оборачиваемость текущих остатков",
                ],
            )

            # Сортируем данные (точно как в оригинале)
            сортировка_артикулов = df.sort_values(
                ["Артикул продавца", "Размер"], ascending=[True, True]
            )

            # СОЗДАЕМ ВЫБОРКУ С СОРТИРОВКОЙ ПО СТАТУСАМ (точно как в оригинале)
            порядок_статусов = [
                "Дефицит",
                "Актуальный",
                "Баланс",
                "Неликвидный",
                "Не рассчитано",
            ]

            def определить_порядок(статус):
                if статус in порядок_статусов:
                    return порядок_статусов.index(статус)
                else:
                    return len(порядок_статусов)

            # Создаем выборку для второй страницы - одна строка на артикул
            выборка_данных = сортировка_артикулов.drop_duplicates(
                subset=["Артикул продавца"]
            )[["Артикул продавца", "Артикул WB", "Доступность"]]

            # Добавляем столбец для сортировки
            выборка_данных["Порядок_сортировки"] = выборка_данных["Доступность"].apply(
                определить_порядок
            )

            # Сортируем по порядку статусов, затем по артикулу
            выборка_данных = выборка_данных.sort_values(
                ["Порядок_сортировки", "Артикул продавца"], ascending=[True, True]
            )

            # Удаляем временный столбец
            выборка_данных = выборка_данных[
                ["Артикул продавца", "Артикул WB", "Доступность"]
            ]

            # Создаем Excel файл с стилями (точно как в оригинале)
            wb = Workbook()
            ws = wb.active
            ws.title = "Отсортированные данные"

            # Записываем данные из DataFrame (точно как в оригинале)
            for r in dataframe_to_rows(сортировка_артикулов, index=False, header=True):
                ws.append(r)

            # Стилизация заголовков (точно как в оригинале)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Цвета для чередования блоков артикулов (точно как в оригинале)
            colors = [
                PatternFill(
                    start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"
                ),  # яркий голубой
                PatternFill(
                    start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"
                ),  # яркий зеленый
                PatternFill(
                    start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
                ),  # яркий оранжевый
                PatternFill(
                    start_color="F8CECC", end_color="F8CECC", fill_type="solid"
                ),  # светло-красный
                PatternFill(
                    start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"
                ),  # светло-фиолетовый
            ]

            # Раскрашиваем блоки артикулов (точно как в оригинале)
            current_article = None
            color_index = 0

            for row in range(
                2, ws.max_row + 1
            ):  # начинаем с 2 строки (после заголовка)
                article = ws.cell(row=row, column=1).value  # столбец "Артикул продавца"

                if article != current_article:
                    current_article = article
                    color_index = (color_index + 1) % len(colors)  # переключаем цвет

                # Применяем цвет ко всей строке
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = colors[color_index]

            # Автоматическая ширина колонок (точно как в оригинале)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

            # СОЗДАЕМ ВТОРУЮ СТРАНИЦУ "Выборка" (точно как в оригинале)
            ws_выборка = wb.create_sheet("Выборка")

            # Записываем данные выборки (точно как в оригинале)
            for r in dataframe_to_rows(выборка_данных, index=False, header=True):
                ws_выборка.append(r)

            # Стилизация заголовков для выборки (точно как в оригинале)
            for cell in ws_выборка[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Цвета для разных статусов на странице Выборка (точно как в оригинале)
            цвета_статусов = {
                "Дефицит": PatternFill(
                    start_color="F8CECC", end_color="F8CECC", fill_type="solid"
                ),  # красный
                "Актуальный": PatternFill(
                    start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"
                ),  # зеленый
                "Баланс": PatternFill(
                    start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"
                ),  # голубой
                "Неликвидный": PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                ),  # желтый
                "Не рассчитано": PatternFill(
                    start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"
                ),  # фиолетовый
            }

            # Цвет по умолчанию для других статусов (точно как в оригинале)
            цвет_по_умолчанию = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            )  # серый

            # Раскрашиваем блоки по статусам на странице Выборка (точно как в оригинале)
            for row in range(2, ws_выборка.max_row + 1):
                status = ws_выборка.cell(
                    row=row, column=3
                ).value  # столбец "Доступность" (3-й столбец)

                # Определяем цвет для статуса
                if status in цвета_статусов:
                    fill_color = цвета_статусов[status]
                else:
                    fill_color = цвет_по_умолчанию

                # Применяем цвет ко всей строке
                for col in range(1, ws_выборка.max_column + 1):
                    ws_выборка.cell(row=row, column=col).fill = fill_color

            # Автоматическая ширина колонок для выборки (точно как в оригинале)
            for column in ws_выборка.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws_выборка.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем во временный файл и возвращаем как FileResponse
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                wb.save(tmp_file.name)

                # Используем FileResponse для автоматической загрузки
                response = FileResponse(
                    open(tmp_file.name, "rb"),
                    as_attachment=True,
                    filename="Признаки_доступности.xlsx",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # Функция для очистки временного файла после отправки
                def cleanup_temp_file():
                    try:
                        os.unlink(tmp_file.name)
                    except:
                        pass

                response.closed = cleanup_temp_file
                return response

        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")

    # Рендерим форму загрузки
    return render(request, "forms_app/form13_simple.html")
