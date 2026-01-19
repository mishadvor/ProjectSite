# forms_app/views/form18_view.py

import pandas as pd
import numpy as np
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from forms_app.models import ArticleCost
from forms_app.forms import ArticleCostForm


def safe_convert_to_int(value):
    try:
        if isinstance(value, (pd.Series, pd.DataFrame)):
            return pd.to_numeric(value, errors="coerce").fillna(0).astype(int)
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def safe_convert_to_float(value):
    try:
        if isinstance(value, (pd.Series, pd.DataFrame)):
            return pd.to_numeric(value, errors="coerce").fillna(0.0)
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def safe_mean_calculation(x, decimals=1):
    try:
        filtered = x[x != 0]
        if len(filtered) > 0:
            return round(float(filtered.mean()), decimals)
        return 0.0
    except Exception:
        return 0.0


@login_required
def form18_list(request):
    if request.method == "POST":
        action = request.POST.get("action", "add_article")

        # === Обработка финансового отчёта ===
        if action == "process_report":
            try:
                file = request.FILES.get("report_file")
                if not file:
                    messages.error(request, "Загрузите файл отчёта.")
                    return redirect("forms_app:form18_list")

                # Параметры из формы
                try:
                    sebestoimost = float(request.POST.get("sebestoimost", 600))
                except (ValueError, TypeError):
                    sebestoimost = 600.0

                try:
                    nalog_str = request.POST.get("nalog_procent", "7").replace(",", ".")
                    nalog_procent = float(nalog_str) / 100
                    nalog_procent = max(0.0, min(1.0, nalog_procent))
                except (ValueError, TypeError):
                    nalog_procent = 0.07

                df = pd.read_excel(file)

                # Приводим Код номенклатуры к строке сразу
                df["Код номенклатуры"] = df["Код номенклатуры"].astype(str).str.strip()

                # Список числовых колонок
                numeric_cols = [
                    "Цена розничная",
                    "Вайлдберриз реализовал Товар (Пр)",
                    "К перечислению Продавцу за реализованный Товар",
                    "Услуги по доставке товара покупателю",
                ]

                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = df[col].apply(safe_convert_to_float)

                # Агрегация по коду номенклатуры
                sums1_per_category = (
                    df.groupby("Код номенклатуры")
                    .agg(
                        {
                            "Артикул поставщика": "first",
                            "Цена розничная": "sum",
                            "Вайлдберриз реализовал Товар (Пр)": "sum",
                            "К перечислению Продавцу за реализованный Товар": "sum",
                            "Услуги по доставке товара покупателю": "sum",
                        }
                    )
                    .reset_index()
                    .query("`Код номенклатуры` != 0")
                )

                for col in numeric_cols:
                    sums1_per_category[col] = sums1_per_category[col].apply(
                        safe_convert_to_int
                    )

                sums1_per_category["К Перечислению без Логистики"] = (
                    sums1_per_category["К перечислению Продавцу за реализованный Товар"]
                    - sums1_per_category["Услуги по доставке товара покупателю"]
                ).apply(safe_convert_to_int)

                sums1_per_category["Сумма СПП"] = (
                    sums1_per_category["Цена розничная"]
                    - sums1_per_category["Вайлдберриз реализовал Товар (Пр)"]
                ).apply(safe_convert_to_int)

                # Расчет процентов с обработкой деления на 0
                sums1_per_category["% Лог/рс"] = (
                    np.where(
                        sums1_per_category[
                            "К перечислению Продавцу за реализованный Товар"
                        ]
                        == 0,
                        0,
                        (
                            sums1_per_category["Услуги по доставке товара покупателю"]
                            / sums1_per_category[
                                "К перечислению Продавцу за реализованный Товар"
                            ]
                        )
                        * 100,
                    )
                ).round(1)

                sums1_per_category["% Лог/Наша Цена"] = (
                    np.where(
                        sums1_per_category["Цена розничная"] == 0,
                        0,
                        (
                            sums1_per_category["Услуги по доставке товара покупателю"]
                            / sums1_per_category["Цена розничная"]
                        )
                        * 100,
                    )
                ).round(1)

                # Возвраты
                returns_by_code = (
                    df[df["Тип документа"] == "Возврат"]
                    .groupby("Код номенклатуры")
                    .agg(
                        {
                            "Цена розничная": "sum",
                            "Вайлдберриз реализовал Товар (Пр)": "sum",
                            "К перечислению Продавцу за реализованный Товар": "sum",
                        }
                    )
                    .reset_index()
                    .rename(
                        columns={
                            "Цена розничная": "Возвраты Наша цена",
                            "Вайлдберриз реализовал Товар (Пр)": "Возвраты реализация ВБ",
                            "К перечислению Продавцу за реализованный Товар": "Возвраты к перечислению",
                        }
                    )
                )

                for col in [
                    "Возвраты Наша цена",
                    "Возвраты реализация ВБ",
                    "Возвраты к перечислению",
                ]:
                    returns_by_code[col] = returns_by_code[col].apply(
                        safe_convert_to_int
                    )

                first_merged = sums1_per_category.merge(
                    returns_by_code, on="Код номенклатуры", how="left"
                ).fillna(0)

                # Расчет чистых продаж
                first_merged["Чистые продажи Наши"] = (
                    first_merged["Цена розничная"] - first_merged["Возвраты Наша цена"]
                ).apply(safe_convert_to_int)

                first_merged["Чистая реализация ВБ"] = (
                    first_merged["Вайлдберриз реализовал Товар (Пр)"]
                    - (first_merged["Возвраты реализация ВБ"] * 2)
                ).apply(safe_convert_to_int)

                first_merged["Чистое Перечисление"] = (
                    first_merged["К перечислению Продавцу за реализованный Товар"]
                    - first_merged["Возвраты к перечислению"]
                ).apply(safe_convert_to_int)

                first_merged["Чистое Перечисление без Логистики"] = (
                    first_merged["Чистое Перечисление"]
                    - first_merged["Услуги по доставке товара покупателю"]
                ).apply(safe_convert_to_int)

                # Средние значения
                cost_per_category = (
                    df.groupby("Код номенклатуры")
                    .agg(
                        {
                            "Артикул поставщика": "first",
                            "Цена розничная": safe_mean_calculation,
                            "Вайлдберриз реализовал Товар (Пр)": safe_mean_calculation,
                            "К перечислению Продавцу за реализованный Товар": safe_mean_calculation,
                            "Услуги по доставке товара покупателю": lambda x: round(
                                safe_convert_to_float(x.mean() * 2), 1
                            ),
                        }
                    )
                    .reset_index()
                )

                cost_per_category["СПП Средняя"] = (
                    cost_per_category["Цена розничная"]
                    - cost_per_category["Вайлдберриз реализовал Товар (Пр)"]
                ).round(1)

                cost_per_category["К Перечислению без Логистики Средняя"] = (
                    cost_per_category["К перечислению Продавцу за реализованный Товар"]
                    - cost_per_category["Услуги по доставке товара покупателю"]
                ).round(1)

                cost_per_category["% Лог/Перечисление с Лог Средний"] = (
                    np.where(
                        cost_per_category[
                            "К перечислению Продавцу за реализованный Товар"
                        ]
                        == 0,
                        0,
                        (
                            cost_per_category["Услуги по доставке товара покупателю"]
                            / cost_per_category[
                                "К перечислению Продавцу за реализованный Товар"
                            ]
                        )
                        * 100,
                    )
                ).round(1)

                cost_per_category["% Лог/Наша цена Средний"] = (
                    np.where(
                        cost_per_category["Цена розничная"] == 0,
                        0,
                        (
                            cost_per_category["Услуги по доставке товара покупателю"]
                            / cost_per_category["Цена розничная"]
                        )
                        * 100,
                    )
                ).round(1)

                second_merged = first_merged.merge(
                    cost_per_category,
                    on="Код номенклатуры",
                    how="left",
                    suffixes=("", "_Среднее"),
                ).fillna(0)

                # === ОБРАБОТКА ЛОГИСТИКИ И СЕБЕСТОИМОСТИ ===
                log_col = next(
                    (col for col in df.columns if "Виды логистики" in col), None
                )
                if log_col:
                    df_exploded = df.explode(log_col)
                    df_exploded[log_col] = df_exploded[log_col].fillna("Не указано")

                    status_log = (
                        df_exploded.groupby("Код номенклатуры")
                        .agg(
                            {
                                log_col: lambda x: x.value_counts().to_dict(),
                                "Артикул поставщика": "first",
                            }
                        )
                        .reset_index()
                    )

                    status_log = pd.concat(
                        [
                            status_log.drop(log_col, axis=1),
                            status_log[log_col].apply(pd.Series).fillna(0),
                        ],
                        axis=1,
                    )

                    required_events = [
                        "К клиенту при продаже",
                        "От клиента при возврате",
                        "От клиента при отмене",
                    ]
                    for event in required_events:
                        if event not in status_log.columns:
                            status_log[event] = 0
                        status_log[event] = (
                            pd.to_numeric(status_log[event], errors="coerce")
                            .fillna(0)
                            .astype(float)
                        )

                    numerator = status_log["К клиенту при продаже"]
                    denominator = (
                        status_log["К клиенту при продаже"]
                        + status_log["От клиента при возврате"]
                        + status_log["От клиента при отмене"]
                    )

                    with np.errstate(divide="ignore", invalid="ignore"):
                        buyout_rate = np.where(
                            denominator == 0, 0.0, (numerator / denominator) * 100
                        )
                    status_log["%Выкупа"] = np.round(buyout_rate.astype(float), 1)

                    # === ИНДИВИДУАЛЬНАЯ СЕБЕСТОИМОСТЬ ИЗ БАЗЫ ===
                    cost_map = {
                        str(ac.wb_article): float(ac.cost)
                        for ac in ArticleCost.objects.filter(user=request.user)
                    }

                    # Себестоимость за 1 шт
                    status_log["Себестоимость за 1 шт"] = (
                        status_log["Код номенклатуры"]
                        .map(cost_map)
                        .fillna(sebestoimost)
                        .astype(float)
                    )

                    # Себестоимость продаж
                    status_log["Себес Продаж"] = (
                        status_log["К клиенту при продаже"]
                        * status_log["Себестоимость за 1 шт"]
                    ).round(0)

                    status_log["Чистые продажи, шт"] = numerator
                    status_log["Заказы"] = denominator

                    third_merged = second_merged.merge(
                        status_log[
                            [
                                "Код номенклатуры",
                                "%Выкупа",
                                "Себес Продаж",
                                "Себестоимость за 1 шт",
                                "Чистые продажи, шт",
                                "Заказы",
                                "От клиента при возврате",
                                "От клиента при отмене",
                            ]
                        ],
                        on="Код номенклатуры",
                        how="left",
                    ).fillna(0)
                else:
                    # Если нет колонки логистики — нельзя определить продажи
                    third_merged = second_merged.copy()
                    third_merged["Себес Продаж"] = 0
                    third_merged["Себестоимость за 1 шт"] = sebestoimost
                    third_merged["Чистые продажи, шт"] = 0
                    third_merged["Заказы"] = 0
                    third_merged["%Выкупа"] = 0.0
                    third_merged["От клиента при возврате"] = 0
                    third_merged["От клиента при отмене"] = 0

                # Переименование столбцов
                third_merged = third_merged.rename(
                    columns={
                        "Цена розничная": "Сумма Продаж Наша Цена",
                        "Вайлдберриз реализовал Товар (Пр)": "Сумма Продаж по цене ВБ",
                        "К перечислению Продавцу за реализованный Товар": "Сумма Продаж Перечисление С Лог",
                        "Услуги по доставке товара покупателю": "Логистика",
                        "От клиента при возврате": "Возвраты, шт",
                        "От клиента при отмене": "Отмена",
                        "Цена розничная_Среднее": "Наша цена Средняя",
                        "Вайлдберриз реализовал Товар (Пр)_Среднее": "Реализация ВБ Средняя",
                        "К перечислению Продавцу за реализованный Товар_Среднее": "К перечислению Среднее",
                        "Услуги по доставке товара покупателю_Среднее": "Логистика Средняя",
                    }
                )

                # =============== ВЫЧИСЛЕНИЕ МАРЖИ И НАЛОГОВ ===============
                third_merged["Маржа"] = (
                    third_merged["Чистое Перечисление без Логистики"]
                    - third_merged["Себес Продаж"]
                ).round(1)

                third_merged["Налоги"] = (
                    third_merged["Чистая реализация ВБ"] * nalog_procent
                ).round(1)

                # =============== ВЫЧИСЛЕНИЕ ДОПОЛНИТЕЛЬНЫХ УДЕРЖАНИЙ ===============
                all_add_log = (
                    df.groupby("Обоснование для оплаты")
                    .agg(
                        {
                            "Услуги по доставке товара покупателю": "sum",
                            "Общая сумма штрафов": "sum",
                            "Хранение": "sum",
                            "Удержания": "sum",
                            "Операции на приемке": "sum",
                        }
                    )
                    .reset_index()
                )

                sum_dop_uderzhany = (
                    all_add_log["Общая сумма штрафов"].sum()
                    + all_add_log["Хранение"].sum()
                    + all_add_log["Удержания"].sum()
                    + all_add_log["Операции на приемке"].sum()
                )

                sum_zakaz = third_merged["Заказы"].sum()

                # Распределяем доп. удержания пропорционально заказам
                third_merged["Доп удержание на кол-во заказов 1 Артикула"] = (
                    (sum_dop_uderzhany / sum_zakaz) * third_merged["Заказы"]
                ).round(1)

                # =============== ПРАВИЛЬНАЯ ПРИБЫЛЬ С УЧЕТОМ ДОП. УДЕРЖАНИЙ ===============
                third_merged["Прибыль"] = (
                    third_merged["Маржа"]
                    - third_merged["Налоги"]
                    - third_merged["Доп удержание на кол-во заказов 1 Артикула"]
                ).round(1)

                # =============== ДОПОЛНИТЕЛЬНЫЕ ПОКАЗАТЕЛИ ===============
                # Прибыль на 1 юбку
                third_merged["Прибыль на 1 Юбку"] = (
                    third_merged.apply(
                        lambda row: (
                            (row["Прибыль"] / row["Чистые продажи, шт"])
                            if row["Чистые продажи, шт"] > 0
                            else (
                                row["Прибыль"] / row["Заказы"]
                                if row["Заказы"] > 0
                                else 0
                            )
                        ),
                        axis=1,
                    )
                    .replace([np.inf, -np.inf], 0)
                    .fillna(0)
                    .round(1)
                )

                # % СПП
                third_merged["% СПП"] = (
                    (third_merged["СПП Средняя"] / third_merged["Наша цена Средняя"])
                    * 100
                ).round(1)

                # Доп. колонки
                third_merged["План на неделю"] = ""
                third_merged["План по доходу"] = ""

                # Порядок колонок
                desired_columns_order = [
                    "Код номенклатуры",
                    "Артикул поставщика",
                    "Чистые продажи Наши",
                    "Чистая реализация ВБ",
                    "Чистое Перечисление",
                    "Чистое Перечисление без Логистики",
                    "Себестоимость за 1 шт",
                    "Себес Продаж",
                    "Прибыль",
                    "Наша цена Средняя",
                    "Реализация ВБ Средняя",
                    "К перечислению Среднее",
                    "Прибыль на 1 Юбку",
                    "Заказы",
                    "Чистые продажи, шт",
                    "%Выкупа",
                    "СПП Средняя",
                    "% СПП",
                    "План на неделю",
                    "План по доходу",
                    "Сумма Продаж Наша Цена",
                    "Сумма Продаж по цене ВБ",
                    "Сумма Продаж Перечисление С Лог",
                    "Логистика",
                    "К Перечислению без Логистики",
                    "Сумма СПП",
                    "% Лог/рс",
                    "% Лог/Наша Цена",
                    "Возвраты Наша цена",
                    "Возвраты реализация ВБ",
                    "Возвраты к перечислению",
                    "Логистика Средняя",
                    "К Перечислению без Логистики Средняя",
                    "% Лог/Перечисление с Лог Средний",
                    "% Лог/Наша цена Средний",
                    "Возвраты, шт",
                    "Отмена",
                    "Маржа",
                    "Налоги",
                    "Доп удержание на кол-во заказов 1 Артикула",
                ]

                existing_columns = [
                    col for col in desired_columns_order if col in third_merged.columns
                ]
                third_merged = third_merged[existing_columns]

                # =============== СОРТИРОВКА ПО ПРИБЫЛИ (УЖЕ С ДОП. УДЕРЖАНИЯМИ) ===============
                third_merged.sort_values(by="Прибыль", ascending=False, inplace=True)

                # =============== ГРУППИРОВКА ПО ПРИБЫЛИ (УЖЕ С ДОП. УДЕРЖАНИЯМИ) ===============
                conditions = [
                    third_merged["Прибыль"] > 10000,
                    (third_merged["Прибыль"] >= 5000)
                    & (third_merged["Прибыль"] <= 10000),
                    (third_merged["Прибыль"] > 0) & (third_merged["Прибыль"] < 5000),
                    third_merged["Прибыль"] < 0,
                ]
                categories_profit = [
                    "1. >10 000",
                    "2. 5 000 - 10 000",
                    "3. 0 - 5 000",
                    "4. <0 (убытки)",
                ]
                third_merged["Группа по прибыли"] = np.select(
                    conditions, categories_profit, default="Не попал"
                )

                # Удаляем строки с нулевой прибылью
                third_merged = third_merged[third_merged["Прибыль"] != 0].copy()

                # =============== ИТОГОВАЯ СВОДКА ===============
                totall_summary = pd.DataFrame(
                    {
                        "Колонка": [
                            "Логистика",
                            "Сумма СПП",
                            "Сумма Чистых перечислений без Возвратов и Логистики",
                            "Чистые продажи, шт",
                            "Заказы",
                            "Себестоимость продаж",
                            "Процент налога",
                            "Налоги",
                            "Прибыль без налога - маржа",
                            "Штрафы",
                            "Хранение",
                            "Удержания",
                            "Операции на приемке",
                            "Прибыль (с учетом доп. удержаний)",
                        ],
                        "Общая сумма": [
                            third_merged["Логистика"].sum(),
                            third_merged["Сумма СПП"].sum(),
                            third_merged["Чистое Перечисление без Логистики"].sum(),
                            third_merged["Чистые продажи, шт"].sum(),
                            third_merged["Заказы"].sum(),
                            third_merged["Себес Продаж"].sum(),
                            f"{round(nalog_procent * 100, 1)}%",
                            third_merged["Налоги"].sum(),
                            third_merged["Маржа"].sum(),
                            all_add_log["Общая сумма штрафов"].sum(),
                            all_add_log["Хранение"].sum(),
                            all_add_log["Удержания"].sum(),
                            all_add_log["Операции на приемке"].sum(),
                            third_merged["Прибыль"].sum(),
                        ],
                    }
                )

                # =============== ГРУППИРОВКА ПО ПРЕФИКСАМ АРТИКУЛОВ ===============
                def get_prefix(article):
                    try:
                        return str(article).split("_")[0][:3]
                    except:
                        return ""

                third_merged["Префикс"] = third_merged["Артикул поставщика"].apply(
                    get_prefix
                )

                # Определяем категории и соответствующие им префиксы артикула
                categories = {
                    "Экокожа черная": ["051", "054", "072", "079", "085", "395"],
                    "Джерси черная": ["001", "002", "003", "004", "005", "050", "122"],
                    "Экокожа цветная": [
                        "052",
                        "053",
                        "056",
                        "057",
                        "058",
                        "059",
                        "060",
                        "061",
                        "062",
                        "063",
                        "064",
                        "065",
                        "066",
                        "067",
                        "068",
                        "069",
                        "070",
                        "071",
                        "073",
                        "074",
                        "075",
                        "076",
                        "077",
                        "078",
                        "080",
                        "081",
                        "082",
                        "083",
                        "084",
                        "093",
                        "100",
                        "102",
                        "103",
                        "123",
                        "101",
                    ],
                    "Джерси цветная": [
                        "006",
                        "007",
                        "008",
                        "009",
                        "010",
                        "011",
                        "012",
                        "013",
                        "014",
                        "015",
                        "016",
                        "017",
                        "018",
                        "019",
                        "020",
                        "021",
                        "022",
                        "023",
                        "024",
                        "025",
                        "026",
                        "027",
                        "028",
                        "029",
                        "030",
                        "031",
                        "032",
                        "033",
                        "034",
                        "035",
                        "036",
                        "037",
                        "038",
                        "039",
                        "040",
                        "041",
                        "042",
                        "043",
                        "044",
                        "045",
                        "046",
                        "047",
                        "048",
                        "049",
                        "055",
                        "086",
                        "087",
                        "088",
                        "089",
                        "090",
                        "091",
                        "092",
                        "094",
                        "095",
                        "096",
                        "097",
                        "098",
                        "099",
                        "104",
                        "105",
                        "106",
                        "107",
                        "108",
                        "109",
                        "110",
                        "111",
                        "112",
                        "113",
                        "114",
                        "115",
                        "116",
                        "117",
                        "118",
                        "119",
                        "120",
                        "121",
                        "131",
                        "132",
                        "133",
                        "281",
                        "341",
                        "342",
                        "343",
                        "344",
                        "345",
                        "346",
                        "347",
                        "348",
                        "349",
                        "350",
                        "351",
                        "354",
                        "355",
                        "356",
                        "357",
                        "358",
                        "387",
                    ],
                    "Джерси Короткая Черная (40,50)": ["352", "353", "392", "395"],
                    "Софт лето": [
                        "203",
                        "197",
                        "206",
                        "168",
                        "169",
                        "170",
                        "171",
                        "172",
                        "173",
                        "174",
                        "175",
                        "176",
                        "177",
                        "178",
                        "179",
                        "180",
                        "181",
                        "182",
                        "183",
                        "184",
                        "185",
                        "186",
                        "187",
                        "188",
                        "189",
                        "190",
                        "191",
                        "192",
                        "193",
                        "194",
                        "195",
                        "196",
                        "197",
                        "198",
                        "199",
                        "200",
                        "201",
                        "202",
                        "203",
                        "204",
                        "205",
                        "206",
                        "207",
                        "208",
                        "209",
                        "210",
                        "211",
                        "212",
                        "213",
                        "225",
                        "226",
                        "227",
                        "228",
                        "229",
                        "230",
                        "232",
                        "233",
                        "234",
                        "235",
                        "236",
                        "237",
                        "238",
                        "239",
                        "240",
                        "241",
                        "242",
                        "243",
                        "244",
                        "245",
                        "246",
                        "247",
                        "248",
                        "249",
                        "250",
                        "251",
                        "252",
                        "253",
                        "254",
                        "255",
                        "256",
                        "258",
                        "259",
                        "260",
                        "261",
                        "262",
                        "263",
                        "264",
                        "265",
                        "266",
                        "267",
                        "268",
                        "269",
                        "270",
                        "271",
                        "272",
                        "273",
                        "274",
                        "275",
                        "276",
                        "277",
                        "278",
                        "279",
                        "280",
                        "282",
                        "283",
                        "284",
                        "285",
                        "286",
                        "287",
                        "288",
                        "289",
                        "290",
                        "291",
                        "292",
                        "293",
                        "294",
                        "295",
                        "296",
                        "297",
                        "298",
                        "299",
                        "300",
                        "301",
                        "302",
                        "303",
                        "304",
                        "305",
                        "306",
                        "307",
                        "308",
                        "309",
                        "310",
                        "311",
                        "312",
                        "313",
                        "314",
                        "315",
                        "317",
                        "318",
                        "319",
                        "320",
                        "321",
                        "322",
                        "323",
                        "324",
                        "325",
                        "326",
                        "328",
                        "329",
                        "330",
                        "331",
                        "332",
                        "333",
                        "334",
                        "335",
                        "336",
                        "337",
                        "338",
                        "340",
                        "359",
                        "360",
                        "361",
                        "362",
                        "363",
                        "364",
                        "365",
                        "366",
                        "367",
                        "368",
                        "369",
                        "370",
                        "371",
                        "372",
                        "373",
                        "374",
                        "375",
                        "376",
                        "377",
                        "378",
                        "379",
                        "380",
                        "381",
                        "382",
                        "383",
                        "385",
                        "386",
                        "224",
                    ],
                    "Другое": [
                        "124",
                        "125",
                        "126",
                        "127",
                        "128",
                        "129",
                        "130",
                        "134",
                        "135",
                        "136",
                        "137",
                        "138",
                        "139",
                        "140",
                        "141",
                        "142",
                        "143",
                        "144",
                        "145",
                        "146",
                        "147",
                        "148",
                        "149",
                        "150",
                        "151",
                        "152",
                        "153",
                        "154",
                        "155",
                        "156",
                        "157",
                        "158",
                        "159",
                        "160",
                        "161",
                        "162",
                        "163",
                        "164",
                        "165",
                        "166",
                        "167",
                        "214",
                        "215",
                        "216",
                        "217",
                        "218",
                        "219",
                        "220",
                        "221",
                        "222",
                        "223",
                        "339",
                    ],
                }

                # =============== ЭКСПОРТ В EXCEL ===============
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    # Лист "Основные данные"
                    third_merged.to_excel(
                        writer,
                        sheet_name="Основные данные",
                        index=False,
                        columns=existing_columns,
                    )

                    # Лист "Итоговая сводка"
                    totall_summary.to_excel(
                        writer, sheet_name="Итоговая сводка", index=False
                    )

                    # Листы с категориями артикулов
                    for category, prefixes in categories.items():
                        filtered = third_merged[
                            third_merged["Префикс"].isin(prefixes)
                        ].drop(columns=["Префикс"])
                        safe_sheet_name = category[:31]
                        filtered.to_excel(
                            writer, sheet_name=safe_sheet_name, index=False
                        )

                        # Специальная обработка для "Джерси Короткая..."
                        if category == "Джерси Короткая Черная (40,50)":
                            ws = writer.sheets[safe_sheet_name]
                            col_names = {
                                col: idx for idx, col in enumerate(filtered.columns, 1)
                            }
                            try:
                                qty_col = col_names["Чистые продажи, шт"]
                                cost_col = col_names["Себес Продаж"]
                                margin_col = col_names["Маржа"]
                                tax_col = col_names["Налоги"]
                                extra_col = col_names[
                                    "Доп удержание на кол-во заказов 1 Артикула"
                                ]
                                profit_col = col_names["Прибыль"]

                                cost_letter = get_column_letter(cost_col)
                                ws[f"{cost_letter}1"] = "Себес Продаж (400р)"

                                for row_idx in range(2, len(filtered) + 2):
                                    qty_cell = f"{get_column_letter(qty_col)}{row_idx}"
                                    cost_cell = (
                                        f"{get_column_letter(cost_col)}{row_idx}"
                                    )
                                    margin_cell = (
                                        f"{get_column_letter(margin_col)}{row_idx}"
                                    )
                                    tax_cell = f"{get_column_letter(tax_col)}{row_idx}"
                                    extra_cell = (
                                        f"{get_column_letter(extra_col)}{row_idx}"
                                    )
                                    profit_cell = (
                                        f"{get_column_letter(profit_col)}{row_idx}"
                                    )

                                    ws[cost_cell] = f"={qty_cell}*400"
                                    ws[margin_cell] = (
                                        f"={get_column_letter(col_names['Чистое Перечисление без Логистики'])}{row_idx}-{cost_cell}"
                                    )
                                    ws[profit_cell] = (
                                        f"={margin_cell}-{tax_cell}-{extra_cell}"
                                    )
                            except KeyError:
                                pass

                    # Листы с группами по прибыли
                    for category in categories_profit:
                        filtered = third_merged[
                            third_merged["Группа по прибыли"] == category
                        ]
                        safe_sheet_name = category[:31]
                        filtered.to_excel(
                            writer, sheet_name=safe_sheet_name, index=False
                        )

                    # Форматирование заголовков
                    header_style = NamedStyle(
                        name="header_style",
                        alignment=Alignment(
                            wrap_text=True, horizontal="center", vertical="center"
                        ),
                        font=Font(bold=True),
                    )
                    for sheet in writer.sheets.values():
                        for cell in sheet[1]:
                            cell.style = header_style
                        for column in sheet.columns:
                            max_length = max(
                                (
                                    len(str(cell.value)) if cell.value else 0
                                    for cell in column[1:]
                                ),
                                default=0,
                            )
                            sheet.column_dimensions[column[0].column_letter].width = (
                                min(max_length + 10, 65)
                            )

                output.seek(0)
                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    'attachment; filename="form18_financial_report.xlsx"'
                )
                return response

            except Exception as e:
                import traceback

                traceback.print_exc()
                messages.error(request, f"Ошибка при обработке: {str(e)}")
                return redirect("forms_app:form18_list")

        # === Добавление артикула ===
        elif action == "add_article":
            form = ArticleCostForm(request.POST)
            if form.is_valid():
                wb_article = form.cleaned_data["wb_article"]
                obj, created = ArticleCost.objects.get_or_create(
                    user=request.user, wb_article=wb_article, defaults=form.cleaned_data
                )
                if created:
                    messages.success(request, "Артикул добавлен!")
                else:
                    for field, value in form.cleaned_data.items():
                        setattr(obj, field, value)
                    obj.save()
                    messages.success(request, "Артикул обновлён!")
            else:
                messages.error(request, "Ошибка в форме.")

    # === GET-запрос ===
    form = ArticleCostForm()
    records = ArticleCost.objects.filter(user=request.user).order_by("-id")
    return render(request, "forms_app/form18.html", {"form": form, "records": records})


@login_required
def form18_edit(request, pk):
    record = get_object_or_404(ArticleCost, pk=pk, user=request.user)
    if request.method == "POST":
        form = ArticleCostForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Запись обновлена!")
            return redirect("forms_app:form18_list")
    else:
        form = ArticleCostForm(instance=record)
    return render(
        request, "forms_app/form18_edit.html", {"form": form, "record": record}
    )


@login_required
def form18_delete(request, pk):
    record = get_object_or_404(ArticleCost, pk=pk, user=request.user)
    record.delete()
    messages.success(request, "Запись удалена.")
    return redirect("forms_app:form18_list")
