# forms_app/views/form10_view.py

import pandas as pd
import numpy as np
from io import BytesIO
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


def form10_view(request):
    result_df = None  # Для отображения в HTML (по умолчанию — по размерам)
    error_message = None

    if request.method == "POST" and request.FILES.get("excel_file"):
        uploaded_file = request.FILES["excel_file"]

        # Проверка расширения файла (игнорируем регистр)
        if not uploaded_file.name.lower().endswith(".xlsx"):
            error_message = "Пожалуйста, загрузите файл в формате .xlsx (поддерживаются .xlsx и .XLSX от Wildberries)."
        else:
            try:
                # Чтение Excel-файла
                df_raw = pd.read_excel(uploaded_file, header=1)
                df_raw = df_raw.reset_index(drop=True)

                # === Лист 1: Статистика по размерам (уже есть) ===
                df1 = (
                    df_raw.groupby(
                        ["Артикул WB", "Баркод", "Артикул продавца", "Размер"],
                        as_index=False,
                    )
                    .agg(
                        {
                            "шт.": "sum",
                            "Сумма заказов минус комиссия WB, руб.": "sum",
                            "Выкупили, шт.": "sum",
                            "К перечислению за товар, руб.": "sum",
                            "Текущий остаток, шт.": "sum",
                        }
                    )
                    .round(0)
                )
                df1 = df1.rename(columns={"шт.": "Заказы, шт."})
                df1 = df1.sort_values(
                    by=["Сумма заказов минус комиссия WB, руб."], ascending=False
                ).reset_index(drop=True)

                # === Лист 2: Статистика по артикулам (без размеров) ===
                df2 = (
                    df_raw.groupby(
                        ["Артикул WB", "Артикул продавца"],
                        as_index=False,
                    )
                    .agg(
                        {
                            "шт.": "sum",
                            "Сумма заказов минус комиссия WB, руб.": "sum",
                            "Выкупили, шт.": "sum",
                            "К перечислению за товар, руб.": "sum",
                            "Текущий остаток, шт.": "sum",
                        }
                    )
                    .round(0)
                )
                df2 = df2.rename(columns={"шт.": "Заказы, шт."})
                df2 = df2.sort_values(
                    by=["Сумма заказов минус комиссия WB, руб."], ascending=False
                ).reset_index(drop=True)

                # === Экспорт в Excel: оба листа ===
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl", mode="w") as writer:
                    # Лист 1: по размерам
                    df1.to_excel(
                        writer, index=False, sheet_name="Стат_продаж_по_размерам"
                    )
                    # Лист 2: по артикулам
                    df2.to_excel(
                        writer, index=False, sheet_name="Стат_продаж_по_артикулам"
                    )

                    workbook = writer.book

                    # Общая функция для форматирования листа
                    def format_worksheet(ws):
                        # Стиль заголовков
                        style_name = "header_style"
                        if style_name not in workbook.named_styles:
                            header_style = NamedStyle(name=style_name)
                            header_style.font = Font(bold=True)
                            header_style.alignment = Alignment(
                                wrap_text=True, horizontal="center", vertical="center"
                            )
                            workbook.add_named_style(header_style)

                        for cell in ws[1]:
                            cell.style = style_name

                        # Автоподбор ширины столбцов
                        for column in ws.columns:
                            max_length = 0
                            col_letter = get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if cell.value not in [None, ""]:
                                        max_length = max(
                                            max_length, len(str(cell.value))
                                        )
                                except:
                                    continue
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[col_letter].width = adjusted_width

                    # Форматируем оба листа
                    format_worksheet(writer.sheets["Стат_продаж_по_размерам"])
                    format_worksheet(writer.sheets["Стат_продаж_по_артикулам"])

                output.seek(0)

                # Возвращаем файл как ответ
                response = HttpResponse(
                    output.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    'attachment; filename="form10_result.xlsx"'
                )
                return response

            except Exception as e:
                error_message = f"Ошибка при обработке файла: {str(e)}"

    # --- Подготовка таблицы для отображения в браузере ---
    # Отображаем df1 (по размерам) в HTML, если нужно
    table_html = (
        result_df.to_html(classes="table table-striped table-bordered", index=False)
        if result_df is not None
        else None
    )

    context = {
        "table_html": table_html,
        "error_message": error_message,
    }
    return render(request, "forms_app/form10.html", context)
