import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import time
import uuid
import warnings
from django import forms
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_protect
from urllib.parse import quote

warnings.filterwarnings("ignore")


# ===== ФОРМА =====
class Form19AdvancedUploadForm(forms.Form):
    """Простая форма загрузки файла для формы 19 - анализ трафика по регионам и городам"""

    file = forms.FileField(
        label="Файл с данными по заказам",
        widget=forms.FileInput(
            attrs={
                "class": "form-control",
                "accept": ".xlsx,.xls,.csv",
            }
        ),
    )

    def clean_file(self):
        file = self.cleaned_data["file"]
        # Проверяем расширение
        if not file.name.endswith((".xlsx", ".xls", ".csv")):
            raise forms.ValidationError(
                "Поддерживаются только файлы Excel (.xlsx, .xls) и CSV (.csv)"
            )
        # Проверяем размер (максимум 50 МБ)
        max_size = 50 * 1024 * 1024  # 50 МБ
        if file.size > max_size:
            raise forms.ValidationError(
                f"Размер файла не должен превышать 50 МБ. Ваш файл: {file.size / (1024*1024):.1f} МБ"
            )
        return file


# ===== VIEW =====
@login_required
@csrf_protect
def form19_view(request):
    """Форма 19 - анализ трафика по регионам и городам"""
    if request.method == "GET":
        # Генерируем новый токен для этой сессии
        request.session["form19_upload_token"] = str(uuid.uuid4())
        request.session.pop("form19_last_upload", None)
        form = Form19AdvancedUploadForm()
        context = {
            "page_title": "Форма 19: Анализ трафика по регионам и городам",
            "upload_token": request.session["form19_upload_token"],
            "form": form,
        }
        return render(request, "forms_app/form19.html", context)

    elif request.method == "POST":
        # Проверяем токен
        post_token = request.POST.get("upload_token")
        session_token = request.session.get("form19_upload_token")
        if not post_token or post_token != session_token:
            messages.warning(
                request, "Неверный токен сессии. Пожалуйста, обновите страницу."
            )
            return redirect("forms_app:form19_view")

        # Проверяем наличие файла
        if "file" not in request.FILES:
            messages.error(request, "❌ Пожалуйста, выберите файл для загрузки!")
            return redirect("forms_app:form19_view")

        try:
            start_time = time.time()
            uploaded_file = request.FILES["file"]

            # Читаем файл
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file, encoding="utf-8")
            else:
                try:
                    df = pd.read_excel(uploaded_file, sheet_name="Заказы", header=1)
                except:
                    df = pd.read_excel(uploaded_file)

            # Определяем колонки
            region_from_col, region_to_col = find_region_columns(df)
            city_from_col, city_to_col = find_city_columns(df)

            # Проверяем, найдены ли какие-либо колонки
            if not (region_from_col and region_to_col) and not (
                city_from_col and city_to_col
            ):
                messages.error(
                    request,
                    "❌ Не удалось определить колонки 'Регион отправки/прибытия' или 'Город отправки/прибытия'. "
                    "Убедитесь, что в файле есть колонки с названиями 'Регион отправки', 'Регион прибытия' "
                    "или неназванные колонки (Unnamed), содержащие слова 'Электросталь', 'Коледино', 'Москва', 'Санкт-Петербург'.",
                )
                return redirect("forms_app:form19_view")

            all_analyses = {}
            all_destination_analyses = {}

            # Анализ по регионам
            if region_from_col and region_to_col:
                print(
                    f"🔍 Найдены колонки регионов: '{region_from_col}' -> '{region_to_col}'"
                )
                region_analysis_result = analyze_traffic(
                    df, region_from_col, region_to_col, "Регионы"
                )
                if region_analysis_result is not None:
                    all_analyses["regions"] = region_analysis_result
                    region_destination_result = analyze_destinations_by_sources(
                        df, region_from_col, region_to_col, "Регионы"
                    )
                    if region_destination_result is not None:
                        all_destination_analyses["regions"] = region_destination_result
                else:
                    print("⚠️ Не удалось провести анализ по регионам.")

            # Анализ по городам
            if city_from_col and city_to_col:
                print(
                    f"🔍 Найдены колонки городов: '{city_from_col}' -> '{city_to_col}'"
                )
                city_analysis_result = analyze_traffic(
                    df, city_from_col, city_to_col, "Города"
                )
                if city_analysis_result is not None:
                    all_analyses["cities"] = city_analysis_result
                    city_destination_result = analyze_destinations_by_sources(
                        df, city_from_col, city_to_col, "Города"
                    )
                    if city_destination_result is not None:
                        all_destination_analyses["cities"] = city_destination_result
                else:
                    print("⚠️ Не удалось провести анализ по городам.")

            # Если не удалось провести ни один анализ
            if not all_analyses:
                messages.error(
                    request,
                    "❌ Ни один из анализов (по регионам или городам) не дал результатов.",
                )
                return redirect("forms_app:form19_view")

            # Создаем Excel отчет
            excel_buffer = create_excel_report_with_proper_names(
                df, all_analyses, all_destination_analyses
            )
            processing_time = time.time() - start_time

            # Сохраняем информацию для статистики (опционально)
            request.session["form19_processing_time"] = processing_time
            request.session["form19_analysis_count"] = len(all_analyses)
            request.session.modified = True

            # Очищаем сессию после отправки файла
            request.session.pop("form19_last_upload", None)

            # Создаем имя файла для скачивания
            original_name = uploaded_file.name.rsplit(".", 1)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"анализ_трафика_результаты_{original_name}_{timestamp}.xlsx"

            # Создаем HttpResponse с правильными заголовками
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            safe_filename = quote(filename.encode("utf-8"))
            response["Content-Disposition"] = (
                f"attachment; filename*=UTF-8''{safe_filename}"
            )
            response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response["Pragma"] = "no-cache"
            response["Expires"] = "Mon, 01 Jan 1990 00:00:00 GMT"

            return response

        except Exception as e:
            messages.error(request, f"❌ Ошибка при обработке файла: {str(e)}")
            import traceback

            print(traceback.format_exc())
            # Возвращаем на главную с новым токеном
            request.session["form19_upload_token"] = str(uuid.uuid4())
            return redirect("forms_app:form19_view")


# ===== ФУНКЦИИ АНАЛИЗА =====
def find_region_columns(df):
    """Находит колонки 'Регион отправки' и 'Регион прибытия' по названию."""
    region_from_col = None
    region_to_col = None

    for col_name in df.columns:
        # Приводим к строке и нижнему регистру для поиска
        name_lower = str(col_name).lower()
        if "регион" in name_lower and "отправ" in name_lower:
            region_from_col = col_name
        elif "регион" in name_lower and (
            "прибыт" in name_lower or "назнач" in name_lower
        ):
            region_to_col = col_name

    return region_from_col, region_to_col


def find_city_columns(df):
    """Находит колонки 'Город отправки' и 'Город прибытия' по ключевым словам в неназванных колонках."""
    city_from_col = None
    city_to_col = None

    # Известные города отправки и прибытия (в нижнем регистре для сравнения)
    known_cities_from_lower = {"электросталь", "коледино"}
    known_cities_to_lower = {
        "москва",
        "санкт-петербург",
        "омск",
        "челябинск",
        "самара",
        "деревня",
        "посёлок",
        "село",
        "тула",
    }

    print("\n🔍 НАЧИНАЕМ ПОИСК КОЛОНОК С ГОРОДАМИ:")
    print(f"Всего колонок в файле: {len(df.columns)}")

    for idx, col_name in enumerate(df.columns):
        col_name_str = str(col_name).strip()
        print(f"\n  Колонка {idx}: '{col_name_str}'")

        # Проверяем, является ли колонка неназванной
        is_unnamed = col_name_str.startswith("Unnamed:") or col_name_str == ""
        print(f"    Тип: {'Неназванная' if is_unnamed else 'С названием'}")

        if not is_unnamed:
            continue  # Пропускаем колонки с названиями

        # Получаем образец данных из этой колонки
        non_null_data = df[col_name].dropna()
        print(f"    Непустых значений: {len(non_null_data)} из {len(df)}")

        if len(non_null_data) == 0:
            print("    Пропускаем - колонка пустая")
            continue

        # Получаем образец данных и приводим к нижнему регистру
        sample_data = non_null_data.head(20).astype(str).str.strip()
        sample_data_lower = sample_data.str.lower().tolist()

        print(
            f"    Образец данных (первые {len(sample_data_lower)}): {sample_data_lower[:5]}"
        )

        # Проверяем, содержит ли колонка известные города отправки
        found_from_cities = []
        for city_word in known_cities_from_lower:
            for data_item in sample_data_lower:
                if city_word in data_item:
                    found_from_cities.append(city_word)
                    break

        if found_from_cities:
            print(f"    Найдены города отправки: {found_from_cities}")
            if city_from_col is None:
                city_from_col = col_name
                print(f"    ✅ Установлена как 'Город отправки': '{col_name}'")
        else:
            # Проверяем, содержит ли колонка известные города прибытия
            found_to_cities = []
            for city_word in known_cities_to_lower:
                for data_item in sample_data_lower:
                    if city_word in data_item:
                        found_to_cities.append(city_word)
                        break

            if found_to_cities:
                print(f"    Найдены города прибытия: {found_to_cities}")
                if city_to_col is None:
                    city_to_col = col_name
                    print(f"    ✅ Установлена как 'Город прибытия': '{col_name}'")
            else:
                print("    Не найдены известные города")

    print(f"\n📊 РЕЗУЛЬТАТЫ ПОИСКА:")
    print(f"  Город отправки: {city_from_col}")
    print(f"  Город прибытия: {city_to_col}")

    return city_from_col, city_to_col


def analyze_traffic(df, from_col, to_col, analysis_name):
    """Основная функция анализа трафика"""
    # Проверяем, что колонки существуют
    if from_col not in df.columns or to_col not in df.columns:
        return None

    # Очистка данных
    df_clean = df.copy()
    df_clean[from_col] = df_clean[from_col].astype(str).str.strip()
    df_clean[to_col] = df_clean[to_col].astype(str).str.strip()

    # Удаляем пустые значения
    df_clean = df_clean[df_clean[from_col] != "nan"]
    df_clean = df_clean[df_clean[to_col] != "nan"]
    df_clean = df_clean[df_clean[from_col] != ""]
    df_clean = df_clean[df_clean[to_col] != ""]

    total_records = len(df_clean)
    if total_records == 0:
        return None

    # 1. Группировка трафика
    traffic_volume = (
        df_clean.groupby([from_col, to_col]).size().reset_index(name="Количество")
    )
    traffic_volume = traffic_volume.sort_values("Количество", ascending=False)

    # 2. Матрица трафика
    traffic_matrix = traffic_volume.pivot_table(
        index=from_col, columns=to_col, values="Количество", fill_value=0
    ).astype(int)

    # 3. Статистика
    unique_sources = df_clean[from_col].nunique()
    unique_destinations = df_clean[to_col].nunique()
    unique_routes = len(traffic_volume)

    # 4. Внутренние/внешние перевозки
    internal = df_clean[df_clean[from_col] == df_clean[to_col]].shape[0]
    external = total_records - internal
    internal_pct = (internal / total_records * 100) if total_records > 0 else 0
    external_pct = (external / total_records * 100) if total_records > 0 else 0

    # 5. Топ маршрутов
    top_n = min(10, len(traffic_volume))
    top_routes = traffic_volume.head(top_n)

    return {
        "df": df_clean,
        "from_col": from_col,
        "to_col": to_col,
        "analysis_name": analysis_name,
        "traffic_volume": traffic_volume,
        "traffic_matrix": traffic_matrix,
        "total_records": total_records,
        "unique_sources": unique_sources,
        "unique_destinations": unique_destinations,
        "unique_routes": unique_routes,
        "internal": internal,
        "external": external,
        "internal_pct": internal_pct,
        "external_pct": external_pct,
        "top_routes": top_routes,
    }


def analyze_destinations_by_sources(df, from_col, to_col, analysis_name):
    """Анализ городов прибытия с группировкой по источникам"""
    # Проверяем, что колонки существуют
    if from_col not in df.columns or to_col not in df.columns:
        return None

    # Очистка данных
    df_clean = df.copy()
    df_clean[from_col] = df_clean[from_col].astype(str).str.strip()
    df_clean[to_col] = df_clean[to_col].astype(str).str.strip()

    # Удаляем пустые значения
    df_clean = df_clean[df_clean[from_col] != "nan"]
    df_clean = df_clean[df_clean[to_col] != "nan"]
    df_clean = df_clean[df_clean[from_col] != ""]
    df_clean = df_clean[df_clean[to_col] != ""]

    total_records = len(df_clean)
    if total_records == 0:
        return None

    # 1. Общее количество по городам прибытия
    destinations_total = (
        df_clean.groupby(to_col).size().reset_index(name="Всего_поступлений")
    )
    destinations_total = destinations_total.sort_values(
        "Всего_поступлений", ascending=False
    )

    # 2. Детализация по источникам для каждого города прибытия
    pivot_table = (
        df_clean.groupby([to_col, from_col]).size().reset_index(name="Количество")
    )
    pivot_table = pivot_table.sort_values(
        [to_col, "Количество"], ascending=[True, False]
    )

    # 3. Создаем структурированные данные для Excel
    destinations_detail = {}
    for dest_city in destinations_total[to_col].head(50):  # Ограничим топ-50 городов
        # Фильтруем данные для конкретного города прибытия
        dest_data = pivot_table[pivot_table[to_col] == dest_city]

        # Сортируем по количеству
        dest_data = dest_data.sort_values("Количество", ascending=False)

        # Рассчитываем процент от общего
        total_to_dest = dest_data["Количество"].sum()
        dest_data["Процент"] = (dest_data["Количество"] / total_to_dest * 100).round(2)

        destinations_detail[dest_city] = {
            "total_received": total_to_dest,
            "sources": dest_data,
            "unique_sources": len(dest_data),
        }

    return {
        "df": df_clean,
        "from_col": from_col,
        "to_col": to_col,
        "analysis_name": analysis_name,
        "destinations_total": destinations_total,
        "destinations_detail": destinations_detail,
        "pivot_table": pivot_table,
        "total_records": total_records,
    }


# ===== ФУНКЦИИ ДЛЯ СОЗДАНИЯ EXCEL ОТЧЕТА С ПРАВИЛЬНЫМИ НАЗВАНИЯМИ =====
def create_excel_report_with_proper_names(df, all_analyses, all_destination_analyses):
    """Создает Excel отчет с правильными названиями из первого кода"""
    from io import BytesIO

    # Создаем рабочую книгу
    wb = Workbook()

    # Удаляем лист по умолчанию
    if "Sheet" in wb.sheetnames:
        ws_default = wb["Sheet"]
        wb.remove(ws_default)

    # 1. Лист с исходными данными
    ws_source = wb.create_sheet("Исходные_данные")
    add_source_data_sheet(ws_source, df, all_analyses)

    # 2. Листы для каждого анализа
    for analysis_type, analysis_data in all_analyses.items():
        analysis_name = analysis_data["analysis_name"]

        # Лист с трафиком между точками
        ws_traffic = wb.create_sheet(f"Трафик_{analysis_name}")
        add_traffic_sheet(ws_traffic, analysis_data)

        # Лист с матрицей трафика (только если не слишком большая)
        if (
            not analysis_data["traffic_matrix"].empty
            and analysis_data["traffic_matrix"].shape[0] <= 30
            and analysis_data["traffic_matrix"].shape[1] <= 30
        ):
            ws_matrix = wb.create_sheet(f"Матрица_{analysis_name}")
            add_traffic_matrix_sheet(ws_matrix, analysis_data)

    # 3. Листы для анализа по городам прибытия (если есть)
    if all_destination_analyses:
        for analysis_type, dest_analysis in all_destination_analyses.items():
            analysis_name = dest_analysis["analysis_name"]

            # Лист с общим количеством по городам прибытия
            ws_dest_summary = wb.create_sheet(f"Место_прибытия_{analysis_name}")
            add_destinations_summary_sheet(ws_dest_summary, dest_analysis)

            # Лист с детализацией источников
            ws_dest_detail = wb.create_sheet(f"Детализация_источников_{analysis_name}")
            add_detailed_sources_sheet(ws_dest_detail, dest_analysis)

    # 4. Лист с информацией об анализе
    ws_info = wb.create_sheet("Информация")
    add_info_sheet(ws_info, all_analyses)

    # 5. Лист со статистикой
    ws_stats = wb.create_sheet("Статистика")
    add_statistics_sheet(ws_stats, all_analyses, all_destination_analyses)

    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def add_source_data_sheet(ws, df, all_analyses):
    """Добавляет лист с исходными данными"""
    # Собираем все колонки для анализа
    columns_to_save = []
    for analysis_type, analysis_data in all_analyses.items():
        from_col = analysis_data["from_col"]
        to_col = analysis_data["to_col"]
        if from_col not in columns_to_save:
            columns_to_save.append(from_col)
        if to_col not in columns_to_save:
            columns_to_save.append(to_col)

    # Добавляем еще несколько важных колонок если они есть
    additional_cols = [
        "Артикул продавца",
        "Название",
        "Бренд",
        "Стоимость",
        "Дата оформления заказа",
    ]
    for col in additional_cols:
        if col in df.columns and col not in columns_to_save:
            columns_to_save.append(col)

    # Записываем заголовки
    ws.append(columns_to_save)

    # Записываем данные (первые 1000 строк для производительности)
    max_rows = min(1000, len(df))
    for idx, row in df.head(max_rows).iterrows():
        row_data = [row[col] if col in df.columns else "" for col in columns_to_save]
        ws.append(row_data)

    # Форматирование
    for i, col in enumerate(columns_to_save, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 25

        for cell in ws[col_letter]:
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_traffic_sheet(ws, analysis_data):
    """Добавляет лист с трафиком между точками"""
    analysis_name = analysis_data["analysis_name"]

    ws.append(
        [
            f"{analysis_name} отправки",
            f"{analysis_name} прибытия",
            "Количество перевозок",
        ]
    )

    # Ограничиваем количество строк для производительности
    max_traffic_rows = min(500, len(analysis_data["traffic_volume"]))
    for _, row in analysis_data["traffic_volume"].head(max_traffic_rows).iterrows():
        ws.append(
            [
                row[analysis_data["from_col"]],
                row[analysis_data["to_col"]],
                row["Количество"],
            ]
        )

    # Форматирование
    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 35
        for cell in ws[col]:
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_traffic_matrix_sheet(ws, analysis_data):
    """Добавляет лист с матрицей трафика"""
    matrix = analysis_data["traffic_matrix"]
    if matrix.empty:
        ws["A1"] = "Нет данных для матрицы"
        return

    # Заголовки
    headers = [f'{analysis_data["analysis_name"]} отправки →'] + list(matrix.columns)
    ws.append(headers)

    # Данные
    for location in matrix.index:
        row_data = [location] + list(matrix.loc[location].values)
        ws.append(row_data)

    # Форматирование
    ws.column_dimensions["A"].width = 35
    for i in range(2, len(matrix.columns) + 2):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 12

    for row in ws.iter_rows(
        min_row=1, max_row=len(matrix) + 1, min_col=1, max_col=len(matrix.columns) + 1
    ):
        for cell in row:
            if cell.row == 1 or cell.column == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )


def add_destinations_summary_sheet(ws, dest_analysis):
    """Добавляет лист с общим количеством по городам прибытия"""
    analysis_name = dest_analysis["analysis_name"]

    ws.append([f"{analysis_name} прибытия", "Всего поступлений", "Процент от общего"])

    for _, row in dest_analysis["destinations_total"].iterrows():
        city = row[dest_analysis["to_col"]]
        total = row["Всего_поступлений"]
        percentage = total / dest_analysis["total_records"] * 100
        ws.append([city, total, f"{percentage:.2f}%"])

    # Форматирование
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )


def add_detailed_sources_sheet(ws, dest_analysis):
    """Добавляет лист с детализацией источников для городов прибытия"""
    analysis_name = dest_analysis["analysis_name"]

    current_row = 1

    # Проходим по топ-30 городам прибытия
    top_destinations = dest_analysis["destinations_total"].head(30)

    for dest_idx, (_, dest_row) in enumerate(top_destinations.iterrows(), 1):
        dest_city = dest_row[dest_analysis["to_col"]]
        total_received = dest_row["Всего_поступлений"]

        # Заголовок для города прибытия
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = (
            f"{dest_idx}. {dest_city} - Всего поступлений: {total_received:,}"
        )
        title_cell.font = Font(bold=True, size=12, color="1F4E79")
        title_cell.fill = PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

        # Заголовки таблицы
        headers = ["№", f"{analysis_name} отправки", "Количество", "Процент"]
        ws.append(headers)

        # Форматирование заголовков
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        current_row += 1

        # Данные по источникам
        if dest_city in dest_analysis["destinations_detail"]:
            sources_data = dest_analysis["destinations_detail"][dest_city]["sources"]

            for src_idx, (_, src_row) in enumerate(
                sources_data.head(20).iterrows(), 1
            ):  # Топ-20 источников
                source_city = src_row[dest_analysis["from_col"]]
                count = src_row["Количество"]
                percent = src_row["Процент"]

                ws.append([src_idx, source_city, count, f"{percent}%"])

                # Форматирование строки
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # Подсветка четных строк
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(
                            start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
                        )

                current_row += 1
        else:
            ws.append(["", "Нет данных", "", ""])
            current_row += 1

        # Добавляем пустую строку между городами
        current_row += 2

    # Настройка ширины колонок
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12


def add_info_sheet(ws, all_analyses):
    """Добавляет лист с информацией об анализе"""
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "ИНФОРМАЦИЯ ОБ АНАЛИЗЕ"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )

    # Информация
    info_data = [
        ["Дата анализа", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    for info_item in info_data:
        ws.append(info_item)

    # Добавляем информацию о каждом анализе
    ws.append(["", ""])
    ws.append(["ВЫПОЛНЕННЫЕ АНАЛИЗЫ:", ""])
    for analysis_type, analysis_data in all_analyses.items():
        ws.append(
            [
                analysis_data["analysis_name"],
                f"{analysis_data['from_col']} → {analysis_data['to_col']}",
            ]
        )

    # Добавляем информацию о дополнительном анализе
    ws.append(["", ""])
    ws.append(["ПРИМЕЧАНИЯ:", ""])
    notes = [
        "1. Анализ выполнен автоматически с определением колонок",
        "2. Данные очищены от пустых значений",
        "3. Топ маршруты отсортированы по количеству перевозок",
        "4. Полные данные доступны в соответствующих листах",
        "5. Для анализа по городам создана детализация по источникам в отдельном листе",
    ]

    for note in notes:
        ws.append([note, ""])

    # Форматирование
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.row == 2:  # Заголовки
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                )

            if cell.row % 2 == 0 and cell.row > 2:
                cell.fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )

            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_statistics_sheet(ws, all_analyses, all_destination_analyses):
    """Добавляет лист со статистикой."""
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "ОБЩАЯ СТАТИСТИКА АНАЛИЗА"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )

    row = 3
    for analysis_type, analysis_data in all_analyses.items():
        analysis_name = analysis_data["analysis_name"]

        # Заголовок раздела
        ws.merge_cells(f"A{row}:C{row}")
        section_cell = ws[f"A{row}"]
        section_cell.value = f"АНАЛИЗ {analysis_name.upper()}"
        section_cell.font = Font(bold=True, size=14, color="2E75B6")
        section_cell.fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        section_cell.alignment = Alignment(horizontal="center")
        row += 1

        # Статистика
        stats_items = [
            ["Колонка отправки", analysis_data["from_col"]],
            ["Колонка прибытия", analysis_data["to_col"]],
            ["Всего перевозок", f"{analysis_data['total_records']:,}"],
            ["Уникальных точек отправки", analysis_data["unique_sources"]],
            ["Уникальных точек прибытия", analysis_data["unique_destinations"]],
            ["Уникальных маршрутов", analysis_data["unique_routes"]],
            [
                "Внутренние перевозки",
                f"{analysis_data['internal']:,} ({analysis_data['internal_pct']:.1f}%)",
            ],
            [
                "Внешние перевозки",
                f"{analysis_data['external']:,} ({analysis_data['external_pct']:.1f}%)",
            ],
        ]
        for stat_name, stat_value in stats_items:
            ws.append([stat_name, stat_value])
            row += 1

        # Топ маршруты
        if not analysis_data["top_routes"].empty:
            ws.append(["", ""])
            ws.append([f"ТОП-10 МАРШРУТОВ ({analysis_name.lower()}):", ""])
            row += 2
            for i in range(min(10, len(analysis_data["top_routes"]))):
                route = analysis_data["top_routes"].iloc[i]
                from_val = route[analysis_data["from_col"]]
                to_val = route[analysis_data["to_col"]]
                count = route["Количество"]
                if len(from_val) > 25:
                    from_val = from_val[:22] + "..."
                if len(to_val) > 25:
                    to_val = to_val[:22] + "..."
                ws.append([f"{i+1}. {from_val} → {to_val}", f"{count:,}"])
                row += 1

        # Статистика по городам/регионам прибытия (если есть)
        if analysis_type in all_destination_analyses:
            dest_analysis = all_destination_analyses[analysis_type]

            # Определяем правильную формулировку в зависимости от типа анализа
            if analysis_name == "Города":
                destination_type = "городам"
                destination_single = "город"
            elif analysis_name == "Регионы":
                destination_type = "регионов"
                destination_single = "регион"
            else:
                destination_type = "точек прибытия"
                destination_single = "точка прибытия"

            ws.append(["", ""])
            ws.append([f"СТАТИСТИКА ПО {destination_type.upper()} ПРИБЫТИЯ:", ""])
            row += 2
            ws.append(
                [
                    f"Всего уникальных {destination_type} прибытия",
                    f"{dest_analysis['destinations_total'].shape[0]}",
                ]
            )
            row += 1
            ws.append([f"ТОП-10 ПО {destination_type.upper()} ПОСТУПЛЕНИЯМ:", ""])
            row += 1
            for i in range(min(10, len(dest_analysis["destinations_total"]))):
                dest_row = dest_analysis["destinations_total"].iloc[i]
                destination = dest_row[dest_analysis["to_col"]]
                total = dest_row["Всего_поступлений"]
                percentage = total / dest_analysis["total_records"] * 100
                if len(destination) > 30:
                    destination = destination[:27] + "..."
                ws.append([f"{i+1}. {destination}", f"{total:,} ({percentage:.1f}%)"])
                row += 1

        row += 2  # Отступ между разделами

    # Настройка ширины колонок
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    # Форматирование
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.row == 1:
                continue
            if cell.value and isinstance(cell.value, str) and ":" in cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="center")
            if cell.row > 2:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
