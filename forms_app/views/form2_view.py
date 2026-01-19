# forms_app/views/form2_view.py
import pandas as pd
import numpy as np
from django.http import HttpResponse
from django.shortcuts import render
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    Border,
    Side,
    PatternFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Side


def safe_convert_to_int(value):
    """Безопасное преобразование в целые числа"""
    try:
        if isinstance(value, (pd.Series, pd.DataFrame)):
            return pd.to_numeric(value, errors="coerce").fillna(0).astype(int)
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def safe_convert_to_float(value):
    """Безопасное преобразование в числа с плавающей точкой"""
    try:
        if isinstance(value, (pd.Series, pd.DataFrame)):
            return pd.to_numeric(value, errors="coerce").fillna(0.0)
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def safe_mean_calculation(x, decimals=1):
    """Безопасный расчет среднего значения"""
    try:
        filtered = x[x != 0]
        if len(filtered) > 0:
            return round(float(filtered.mean()), decimals)
        return 0.0
    except Exception:
        return 0.0


def form2(request):
    if request.method == "POST":
        mode = request.POST.get("mode")
        # Получаем себестоимость из формы
        try:
            sebestoimost = float(request.POST.get("sebestoimost", 600))
        except (ValueError, TypeError):
            sebestoimost = 600.0
        # Получаем процент налога из формы
        try:
            # Получаем процент как строку и заменяем запятую на точку
            nalog_procent_str = request.POST.get("nalog_procent", "7")
            nalog_procent_str = nalog_procent_str.replace(",", ".")
            # Преобразуем в float и делим на 100 для получения коэффициента
            nalog_procent = float(nalog_procent_str) / 100
            # Ограничиваем значение от 0 до 1
            nalog_procent = max(0.0, min(1.0, nalog_procent))
        except (ValueError, TypeError):
            nalog_procent = 0.07  # Значение по умолчанию 7%
        try:
            # Загрузка файлов
            if mode == "single":
                file = request.FILES.get("file_single")
                if not file:
                    return render(
                        request,
                        "forms_app/form2.html",
                        {"error": "Необходимо загрузить файл."},
                    )

                df = pd.read_excel(file)
            elif mode == "combined":
                file_russia = request.FILES.get("file_russia")
                file_cis = request.FILES.get("file_cis")
                if not file_russia or not file_cis:
                    return render(
                        request,
                        "forms_app/form2.html",
                        {"error": "Пожалуйста, загрузите оба файла."},
                    )

                df_russia = pd.read_excel(file_russia)
                df_cis = pd.read_excel(file_cis)
                df = pd.concat([df_russia, df_cis], ignore_index=True)
            else:
                return render(
                    request, "forms_app/form2.html", {"error": "Неизвестный режим."}
                )

            # Список числовых колонок для обработки
            numeric_cols = [
                "Цена розничная",
                "Вайлдберриз реализовал Товар (Пр)",
                "К перечислению Продавцу за реализованный Товар",
                "Услуги по доставке товара покупателю",
            ]

            # Предварительная обработка числовых колонок
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = df[col].apply(safe_convert_to_float)

            # Основная агрегация по коду номенклатуры
            sums1_per_category = (
                df.groupby("Код номенклатуры")
                .agg(
                    {
                        "Артикул поставщика": "first",
                        "Цена розничная": "sum",
                        "Вайлдберриз реализовал Товар (Пр)": "sum",
                        "К перечислению Продавцу за реализованный Товар": "sum",
                        "Услуги по доставке товара покупателю": "sum",
                    }
                )
                .reset_index()
                # Удаляем строки где код номенклатуры равен 0
                .query("`Код номенклатуры` != 0")
            )

            # Безопасное преобразование числовых колонок
            for col in numeric_cols:
                sums1_per_category[col] = sums1_per_category[col].apply(
                    safe_convert_to_int
                )

            # Дополнительные расчеты
            sums1_per_category["К Перечислению без Логистики"] = (
                sums1_per_category["К перечислению Продавцу за реализованный Товар"]
                - sums1_per_category["Услуги по доставке товара покупателю"]
            ).apply(safe_convert_to_int)

            sums1_per_category["Сумма СПП"] = (
                sums1_per_category["Цена розничная"]
                - sums1_per_category["Вайлдберриз реализовал Товар (Пр)"]
            ).apply(safe_convert_to_int)

            # Расчет процентов с обработкой деления на 0
            sums1_per_category["% Лог/рс"] = (
                np.where(
                    sums1_per_category["К перечислению Продавцу за реализованный Товар"]
                    == 0,
                    0,
                    (
                        sums1_per_category["Услуги по доставке товара покупателю"]
                        / sums1_per_category[
                            "К перечислению Продавцу за реализованный Товар"
                        ]
                    )
                    * 100,
                )
            ).round(1)

            sums1_per_category["% Лог/Наша Цена"] = (
                np.where(
                    sums1_per_category["Цена розничная"] == 0,
                    0,
                    (
                        sums1_per_category["Услуги по доставке товара покупателю"]
                        / sums1_per_category["Цена розничная"]
                    )
                    * 100,
                )
            ).round(1)

            # Агрегация возвратов по коду номенклатуры
            returns_by_code = (
                df[df["Тип документа"] == "Возврат"]
                .groupby("Код номенклатуры")
                .agg(
                    {
                        "Цена розничная": "sum",
                        "Вайлдберриз реализовал Товар (Пр)": "sum",
                        "К перечислению Продавцу за реализованный Товар": "sum",
                    }
                )
                .reset_index()
                .rename(
                    columns={
                        "Цена розничная": "Возвраты Наша цена",
                        "Вайлдберриз реализовал Товар (Пр)": "Возвраты реализация ВБ",
                        "К перечислению Продавцу за реализованный Товар": "Возвраты к перечислению",
                    }
                )
            )

            # Безопасное преобразование возвратов
            for col in [
                "Возвраты Наша цена",
                "Возвраты реализация ВБ",
                "Возвраты к перечислению",
            ]:
                returns_by_code[col] = returns_by_code[col].apply(safe_convert_to_int)

            # Объединение данных
            first_merged = sums1_per_category.merge(
                returns_by_code, on="Код номенклатуры", how="left"
            ).fillna(0)

            # Расчет чистых продаж
            first_merged["Чистые продажи Наши"] = (
                first_merged["Цена розничная"] - first_merged["Возвраты Наша цена"]
            ).apply(safe_convert_to_int)

            first_merged["Чистая реализация ВБ"] = (
                first_merged["Вайлдберриз реализовал Товар (Пр)"]
                - (first_merged["Возвраты реализация ВБ"] * 2)
            ).apply(safe_convert_to_int)

            first_merged["Чистое Перечисление"] = (
                first_merged["К перечислению Продавцу за реализованный Товар"]
                - first_merged["Возвраты к перечислению"]
            ).apply(safe_convert_to_int)

            first_merged["Чистое Перечисление без Логистики"] = (
                first_merged["Чистое Перечисление"]
                - first_merged["Услуги по доставке товара покупателю"]
            ).apply(safe_convert_to_int)

            # Расчет средних значений по коду номенклатуры
            # =====================================================
            cost_per_category = (
                df.groupby("Код номенклатуры")
                .agg(
                    {
                        "Артикул поставщика": "first",
                        "Цена розничная": safe_mean_calculation,
                        "Вайлдберриз реализовал Товар (Пр)": safe_mean_calculation,
                        "К перечислению Продавцу за реализованный Товар": safe_mean_calculation,
                        "Услуги по доставке товара покупателю": lambda x: round(
                            safe_convert_to_float(x.mean() * 2), 1
                        ),
                    }
                )
                .reset_index()
            )

            # Дополнительные расчеты для средних значений
            cost_per_category["СПП Средняя"] = (
                cost_per_category["Цена розничная"]
                - cost_per_category["Вайлдберриз реализовал Товар (Пр)"]
            ).round(1)

            cost_per_category["К Перечислению без Логистики Средняя"] = (
                cost_per_category["К перечислению Продавцу за реализованный Товар"]
                - cost_per_category["Услуги по доставке товара покупателю"]
            ).round(1)

            cost_per_category["% Лог/Перечисление с Лог Средний"] = (
                np.where(
                    cost_per_category["К перечислению Продавцу за реализованный Товар"]
                    == 0,
                    0,
                    (
                        cost_per_category["Услуги по доставке товара покупателю"]
                        / cost_per_category[
                            "К перечислению Продавцу за реализованный Товар"
                        ]
                    )
                    * 100,
                )
            ).round(1)

            cost_per_category["% Лог/Наша цена Средний"] = (
                np.where(
                    cost_per_category["Цена розничная"] == 0,
                    0,
                    (
                        cost_per_category["Услуги по доставке товара покупателю"]
                        / cost_per_category["Цена розничная"]
                    )
                    * 100,
                )
            ).round(1)

            # Объединение со средними значениями
            second_merged = first_merged.merge(
                cost_per_category,
                on="Код номенклатуры",
                how="left",
                suffixes=("", "_Среднее"),
            ).fillna(0)

            # Обработка логистики (если есть соответствующая колонка)
            log_col = next((col for col in df.columns if "Виды логистики" in col), None)
            if log_col:
                df_exploded = df.explode(log_col)
                df_exploded[log_col] = df_exploded[log_col].fillna("Не указано")

                status_log = (
                    df_exploded.groupby("Код номенклатуры")
                    .agg(
                        {
                            log_col: lambda x: x.value_counts().to_dict(),
                            "Артикул поставщика": "first",
                        }
                    )
                    .reset_index()
                )

                # Раскрываем словарь в колонки
                status_log = pd.concat(
                    [
                        status_log.drop(log_col, axis=1),
                        status_log[log_col].apply(pd.Series).fillna(0),
                    ],
                    axis=1,
                )

                # Убедимся, что нужные колонки существуют и числовые
                required_events = [
                    "К клиенту при продаже",
                    "От клиента при возврате",
                    "От клиента при отмене",
                ]
                for event in required_events:
                    if event not in status_log.columns:
                        status_log[event] = 0
                    # Приводим к числу (на случай, если остались строки)
                    status_log[event] = (
                        pd.to_numeric(status_log[event], errors="coerce")
                        .fillna(0)
                        .astype(float)
                    )

                # Расчёт числителя и знаменателя
                numerator = status_log["К клиенту при продаже"]
                denominator = (
                    status_log["К клиенту при продаже"]
                    + status_log["От клиента при возврате"]
                    + status_log["От клиента при отмене"]
                )

                # Безопасный расчёт %Выкупа без .fillna() на ndarray
                with np.errstate(divide="ignore", invalid="ignore"):
                    buyout_rate = np.where(
                        denominator == 0, 0.0, (numerator / denominator) * 100
                    )

                # Присваиваем результат (округляем до 1 знака)
                status_log["%Выкупа"] = np.round(buyout_rate.astype(float), 1)

                # Дополнительные расчёты
                status_log["Себес Продаж"] = (numerator * sebestoimost).round(0)
                status_log["Чистые продажи, шт"] = numerator
                status_log["Заказы"] = denominator

                # Объединение с данными логистики
                third_merged = second_merged.merge(
                    status_log[
                        [
                            "Код номенклатуры",
                            "%Выкупа",
                            "Себес Продаж",
                            "Чистые продажи, шт",
                            "Заказы",
                            "От клиента при возврате",
                            "От клиента при отмене",
                        ]
                    ],
                    on="Код номенклатуры",
                    how="left",
                ).fillna(0)
            else:
                third_merged = second_merged

            # Переименование столбцов
            third_merged = third_merged.rename(
                columns={
                    "Цена розничная": "Сумма Продаж Наша Цена",
                    "Вайлдберриз реализовал Товар (Пр)": "Сумма Продаж по цене ВБ",
                    "К перечислению Продавцу за реализованный Товар": "Сумма Продаж Перечисление С Лог",
                    "Услуги по доставке товара покупателю": "Логистика",
                    "От клиента при возврате": "Возвраты, шт",
                    "От клиента при отмене": "Отмена",
                }
            )

            # =============== ФИНАЛЬНЫЕ РАСЧЕТЫ ===============
            # Маржа
            third_merged["Маржа"] = (
                third_merged["Чистое Перечисление без Логистики"]
                - third_merged["Себес Продаж"]
            ).round(1)

            # Налоги
            third_merged["Налоги"] = (
                third_merged["Чистая реализация ВБ"] * nalog_procent
            ).round(1)

            # =============== ВЫЧИСЛЯЕМ ДОПОЛНИТЕЛЬНЫЕ УДЕРЖАНИЯ ===============
            all_add_log = (
                df.groupby("Обоснование для оплаты")
                .agg(
                    {
                        "Услуги по доставке товара покупателю": "sum",
                        "Общая сумма штрафов": "sum",
                        "Хранение": "sum",
                        "Удержания": "sum",
                        "Операции на приемке": "sum",
                    }
                )
                .reset_index()
            )

            sum_dop_uderzhany = (
                all_add_log["Общая сумма штрафов"].sum()
                + all_add_log["Хранение"].sum()
                + all_add_log["Удержания"].sum()
                + all_add_log["Операции на приемке"].sum()
            )

            sum_zakaz = third_merged["Заказы"].sum()

            # Распределяем доп. удержания пропорционально заказам
            third_merged["Доп удержание на кол-во заказов 1 Артикула"] = (
                (sum_dop_uderzhany / sum_zakaz) * third_merged["Заказы"]
            ).round(1)

            # =============== ПРАВИЛЬНАЯ ПРИБЫЛЬ С УЧЕТОМ ДОП. УДЕРЖАНИЙ ===============
            third_merged["Прибыль"] = (
                third_merged["Маржа"]
                - third_merged["Налоги"]
                - third_merged["Доп удержание на кол-во заказов 1 Артикула"]
            ).round(1)

            # Прибыль на 1 юбку
            third_merged["Прибыль на 1 Юбку"] = (
                third_merged.apply(
                    lambda row: (
                        (row["Прибыль"] / row["Чистые продажи, шт"])
                        if row["Чистые продажи, шт"] > 0
                        else (
                            (row["Прибыль"] / row["Заказы"]) if row["Заказы"] > 0 else 0
                        )
                    ),
                    axis=1,
                )
                .replace([np.inf, -np.inf], 0)
                .fillna(0)
                .round(1)
            )

            # Добавляем пустые колонки перед определением порядка
            third_merged["План на неделю"] = ""
            third_merged["План по доходу"] = ""
            third_merged["% СПП"] = ""

            third_merged = third_merged.rename(
                columns={
                    "Цена розничная": "Сумма Продаж Наша Цена",
                    "Вайлдберриз реализовал Товар (Пр)": "Сумма Продаж по цене ВБ",
                    "К перечислению Продавцу за реализованный Товар": "Сумма Продаж Перечисление С Лог",
                    "Услуги по доставке товара покупателю": "Логистика",
                    "От клиента при возврате": "Возвраты, шт",
                    "От клиента при отмене": "Отмена",
                    "Цена розничная_Среднее": "Наша цена Средняя",
                    "Вайлдберриз реализовал Товар (Пр)_Среднее": "Реализация ВБ Средняя",
                    "К перечислению Продавцу за реализованный Товар_Среднее": "К перечислению Среднее",
                    "Услуги по доставке товара покупателю_Среднее": "Логистика Средняя",
                }
            )

            # Определяем желаемый порядок колонок
            desired_columns_order = [
                "Код номенклатуры",
                "Артикул поставщика",
                "Чистые продажи Наши",
                "Чистая реализация ВБ",
                "Чистое Перечисление",
                "Чистое Перечисление без Логистики",
                "Себес Продаж",
                "Прибыль",
                "Наша цена Средняя",
                "Реализация ВБ Средняя",
                "К перечислению Среднее",
                "Прибыль на 1 Юбку",
                "Заказы",
                "Чистые продажи, шт",
                "%Выкупа",
                "СПП Средняя",
                "% СПП",
                "План на неделю",
                "План по доходу",
                "Сумма Продаж Наша Цена",
                "Сумма Продаж по цене ВБ",
                "Сумма Продаж Перечисление С Лог",
                "Логистика",
                "К Перечислению без Логистики",
                "Сумма СПП",
                "% Лог/рс",
                "% Лог/Наша Цена",
                "Возвраты Наша цена",
                "Возвраты реализация ВБ",
                "Возвраты к перечислению",
                "Логистика Средняя",
                "К Перечислу без Логистики Средняя",
                "% Лог/Перечисление с Лог Средний",
                "% Лог/Наша цена Средний",
                "Возвраты, шт",
                "Отмена",
                "Маржа",
                "Налоги",
                "Доп удержание на кол-во заказов 1 Артикула",
            ]

            # Фильтруем только существующие колонки
            existing_columns = [
                col for col in desired_columns_order if col in third_merged.columns
            ]
            third_merged = third_merged[existing_columns]

            # =============== ПЕРЕСЧЕТ % СПП ===============
            third_merged["% СПП"] = (
                (third_merged["СПП Средняя"] / third_merged["Наша цена Средняя"]) * 100
            ).round(1)

            # =============== СОРТИРОВКА ПО ПРИБЫЛИ (УЖЕ С ДОП. УДЕРЖАНИЯМИ) ===============
            third_merged.sort_values(by="Прибыль", ascending=False, inplace=True)

            # =============== ГРУППИРОВКА ПО ПРИБЫЛИ (УЖЕ С ДОП. УДЕРЖАНИЯМИ) ===============
            conditions = [
                third_merged["Прибыль"] > 10000,
                (third_merged["Прибыль"] >= 5000) & (third_merged["Прибыль"] <= 10000),
                (third_merged["Прибыль"] > 0) & (third_merged["Прибыль"] < 5000),
                third_merged["Прибыль"] < 0,
            ]

            categories_profit = [
                "1. >10 000",
                "2. 5 000 - 10 000",
                "3. 0 - 5 000",
                "4. <0 (убытки)",
            ]

            third_merged["Группа по прибыли"] = np.select(
                conditions, categories_profit, default="Не попал"
            )

            # =============== ИТОГОВАЯ СВОДКА ===============
            totall_summary = pd.DataFrame(
                {
                    "Колонка": [
                        "Логистика",
                        "Сумма СПП",
                        "Сумма Чистых перечислений без Возвратов и Логистики",
                        "Чистые продажи, шт",
                        "Заказы",
                        "Себестоимость продаж",
                        "Процент налога",
                        "Налоги",
                        "Прибыль без налога - маржа",
                        "Штрафы",
                        "Хранение",
                        "Удержания",
                        "Операции на приемке",
                        "Прибыль (с учетом доп. удержаний)",
                    ],
                    "Общая сумма": [
                        third_merged["Логистика"].sum(),
                        third_merged["Сумма СПП"].sum(),
                        third_merged["Чистое Перечисление без Логистики"].sum(),
                        third_merged["Чистые продажи, шт"].sum(),
                        third_merged["Заказы"].sum(),
                        third_merged["Себес Продаж"].sum(),
                        f"{round(nalog_procent * 100, 1)}%",
                        third_merged["Налоги"].sum(),
                        third_merged["Маржа"].sum(),
                        all_add_log["Общая сумма штрафов"].sum(),
                        all_add_log["Хранение"].sum(),
                        all_add_log["Удержания"].sum(),
                        all_add_log["Операции на приемке"].sum(),
                        third_merged["Прибыль"].sum(),
                    ],
                }
            )

            # Удаляем строки с нулевой прибылью
            third_merged = third_merged[third_merged["Прибыль"] != 0].copy()

            # =============== ГРУППИРОВКА ПО ПРЕФИКСАМ АРТИКУЛОВ ===============
            def get_prefix(article):
                """Извлечение первых трёх символов артикула"""
                try:
                    return str(article).split("_")[0][:3]
                except:
                    return ""

            third_merged["Префикс"] = third_merged["Артикул поставщика"].apply(
                get_prefix
            )

            # Определяем категории и соответствующие им префиксы артикула
            categories = {
                "Экокожа черная": ["051", "054", "072", "079", "085", "395"],
                "Джерси черная": ["001", "002", "003", "004", "005", "050", "122"],
                "Экокожа цветная": [
                    "052",
                    "053",
                    "056",
                    "057",
                    "058",
                    "059",
                    "060",
                    "061",
                    "062",
                    "063",
                    "064",
                    "065",
                    "066",
                    "067",
                    "068",
                    "069",
                    "070",
                    "071",
                    "073",
                    "074",
                    "075",
                    "076",
                    "077",
                    "078",
                    "080",
                    "081",
                    "082",
                    "083",
                    "084",
                    "093",
                    "100",
                    "102",
                    "103",
                    "123",
                    "101",
                ],
                "Джерси цветная": [
                    "006",
                    "007",
                    "008",
                    "009",
                    "010",
                    "011",
                    "012",
                    "013",
                    "014",
                    "015",
                    "016",
                    "017",
                    "018",
                    "019",
                    "020",
                    "021",
                    "022",
                    "023",
                    "024",
                    "025",
                    "026",
                    "027",
                    "028",
                    "029",
                    "030",
                    "031",
                    "032",
                    "033",
                    "034",
                    "035",
                    "036",
                    "037",
                    "038",
                    "039",
                    "040",
                    "041",
                    "042",
                    "043",
                    "044",
                    "045",
                    "046",
                    "047",
                    "048",
                    "049",
                    "055",
                    "086",
                    "087",
                    "088",
                    "089",
                    "090",
                    "091",
                    "092",
                    "094",
                    "095",
                    "096",
                    "097",
                    "098",
                    "099",
                    "104",
                    "105",
                    "106",
                    "107",
                    "108",
                    "109",
                    "110",
                    "111",
                    "112",
                    "113",
                    "114",
                    "115",
                    "116",
                    "117",
                    "118",
                    "119",
                    "120",
                    "121",
                    "131",
                    "132",
                    "133",
                    "281",
                    "341",
                    "342",
                    "343",
                    "344",
                    "345",
                    "346",
                    "347",
                    "348",
                    "349",
                    "350",
                    "351",
                    "354",
                    "355",
                    "356",
                    "357",
                    "358",
                    "387",
                ],
                "Джерси Короткая Черная (40,50)": ["352", "353", "392", "395"],
                "Софт лето": [
                    "203",
                    "197",
                    "206",
                    "168",
                    "169",
                    "170",
                    "171",
                    "172",
                    "173",
                    "174",
                    "175",
                    "176",
                    "177",
                    "178",
                    "179",
                    "180",
                    "181",
                    "182",
                    "183",
                    "184",
                    "185",
                    "186",
                    "187",
                    "188",
                    "189",
                    "190",
                    "191",
                    "192",
                    "193",
                    "194",
                    "195",
                    "196",
                    "197",
                    "198",
                    "199",
                    "200",
                    "201",
                    "202",
                    "203",
                    "204",
                    "205",
                    "206",
                    "207",
                    "208",
                    "209",
                    "210",
                    "211",
                    "212",
                    "213",
                    "225",
                    "226",
                    "227",
                    "228",
                    "229",
                    "230",
                    "232",
                    "233",
                    "234",
                    "235",
                    "236",
                    "237",
                    "238",
                    "239",
                    "240",
                    "241",
                    "242",
                    "243",
                    "244",
                    "245",
                    "246",
                    "247",
                    "248",
                    "249",
                    "250",
                    "251",
                    "252",
                    "253",
                    "254",
                    "255",
                    "256",
                    "258",
                    "259",
                    "260",
                    "261",
                    "262",
                    "263",
                    "264",
                    "265",
                    "266",
                    "267",
                    "268",
                    "269",
                    "270",
                    "271",
                    "272",
                    "273",
                    "274",
                    "275",
                    "276",
                    "277",
                    "278",
                    "279",
                    "280",
                    "282",
                    "283",
                    "284",
                    "285",
                    "286",
                    "287",
                    "288",
                    "289",
                    "290",
                    "291",
                    "292",
                    "293",
                    "294",
                    "295",
                    "296",
                    "297",
                    "298",
                    "299",
                    "300",
                    "301",
                    "302",
                    "303",
                    "304",
                    "305",
                    "306",
                    "307",
                    "308",
                    "309",
                    "310",
                    "311",
                    "312",
                    "313",
                    "314",
                    "315",
                    "317",
                    "318",
                    "319",
                    "320",
                    "321",
                    "322",
                    "323",
                    "324",
                    "325",
                    "326",
                    "328",
                    "329",
                    "330",
                    "331",
                    "332",
                    "333",
                    "334",
                    "335",
                    "336",
                    "337",
                    "338",
                    "340",
                    "359",
                    "360",
                    "361",
                    "362",
                    "363",
                    "364",
                    "365",
                    "366",
                    "367",
                    "368",
                    "369",
                    "370",
                    "371",
                    "372",
                    "373",
                    "374",
                    "375",
                    "376",
                    "377",
                    "378",
                    "379",
                    "380",
                    "381",
                    "382",
                    "383",
                    "385",
                    "386",
                    "224",
                ],
                "Другое": [
                    "124",
                    "125",
                    "126",
                    "127",
                    "128",
                    "129",
                    "130",
                    "134",
                    "135",
                    "136",
                    "137",
                    "138",
                    "139",
                    "140",
                    "141",
                    "142",
                    "143",
                    "144",
                    "145",
                    "146",
                    "147",
                    "148",
                    "149",
                    "150",
                    "151",
                    "152",
                    "153",
                    "154",
                    "155",
                    "156",
                    "157",
                    "158",
                    "159",
                    "160",
                    "161",
                    "162",
                    "163",
                    "164",
                    "165",
                    "166",
                    "167",
                    "214",
                    "215",
                    "216",
                    "217",
                    "218",
                    "219",
                    "220",
                    "221",
                    "222",
                    "223",
                    "339",
                ],
            }

            # Генерация Excel файла
            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Лист "Основные данные"
                third_merged.to_excel(
                    writer,
                    sheet_name="Основные данные",
                    index=False,
                    columns=existing_columns,
                )

                # Лист "Итоговая сводка"
                totall_summary.to_excel(
                    writer, sheet_name="Итоговая сводка", index=False
                )

                # Листы с категориями артикулов
                for category, prefixes in categories.items():
                    filtered = third_merged[
                        third_merged["Префикс"].isin(prefixes)
                    ].drop(columns=["Префикс"])
                    safe_sheet_name = category[:31]
                    filtered.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                    # Специальная обработка для "Джерси Короткая Черная (40,50)"
                    if category == "Джерси Короткая Черная (40,50)":
                        ws = writer.sheets[safe_sheet_name]
                        col_names = {
                            col: idx for idx, col in enumerate(filtered.columns, 1)
                        }
                        try:
                            qty_col = col_names["Чистые продажи, шт"]
                            cost_col = col_names["Себес Продаж"]
                            margin_col = col_names["Маржа"]
                            tax_col = col_names["Налоги"]
                            extra_deduction_col = col_names[
                                "Доп удержание на кол-во заказов 1 Артикула"
                            ]
                            profit_col = col_names["Прибыль"]

                            cost_col_letter = get_column_letter(cost_col)
                            ws[f"{cost_col_letter}1"] = "Себес Продаж (400р)"

                            for row_idx in range(2, len(filtered) + 2):
                                qty_cell = f"{get_column_letter(qty_col)}{row_idx}"
                                cost_cell = f"{get_column_letter(cost_col)}{row_idx}"
                                ws[cost_cell] = f"={qty_cell}*400"

                                clean_payment_col = col_names[
                                    "Чистое Перечисление без Логистики"
                                ]
                                clean_payment_cell = (
                                    f"{get_column_letter(clean_payment_col)}{row_idx}"
                                )
                                margin_cell = (
                                    f"{get_column_letter(margin_col)}{row_idx}"
                                )
                                ws[margin_cell] = f"={clean_payment_cell}-{cost_cell}"

                                tax_cell = f"{get_column_letter(tax_col)}{row_idx}"
                                extra_cell = (
                                    f"{get_column_letter(extra_deduction_col)}{row_idx}"
                                )
                                profit_cell = (
                                    f"{get_column_letter(profit_col)}{row_idx}"
                                )
                                ws[profit_cell] = (
                                    f"={margin_cell}-{tax_cell}-{extra_cell}"
                                )

                        except KeyError as e:
                            print(f"Не хватает колонки для пересчёта: {e}")

                # Листы с группами по прибыли
                for category in categories_profit:
                    filtered = third_merged[
                        third_merged["Группа по прибыли"] == category
                    ]
                    safe_sheet_name = category[:31]
                    filtered.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                # Форматирование заголовков
                header_style = NamedStyle(
                    name="header_style",
                    alignment=Alignment(
                        wrap_text=True, horizontal="center", vertical="center"
                    ),
                    font=Font(bold=True),
                )

                for sheetname in writer.sheets:
                    sheet = writer.sheets[sheetname]

                    for cell in sheet[1]:
                        cell.style = header_style

                    for column in sheet.columns:
                        max_length = max(
                            (
                                len(str(cell.value)) if cell.value else 0
                                for cell in column[1:]
                            ),
                            default=0,
                        )
                        sheet.column_dimensions[column[0].column_letter].width = min(
                            max_length + 10, 65
                        )

            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="wildberries_report_Form_2.xlsx"'
            )
            return response

        except Exception as e:
            import traceback

            traceback.print_exc()
            return render(
                request,
                "forms_app/form2.html",
                {"error": f"Ошибка при обработке: {str(e)}"},
            )

    return render(request, "forms_app/form2.html")
