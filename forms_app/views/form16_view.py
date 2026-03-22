# forms_app/views/form16_view.py

import pandas as pd
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from django.contrib import messages
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io
import os
import tempfile

# Импортируем модель (она теперь определена в models.py)
from forms_app.models import Form16Article
from forms_app.forms import Form16UploadForm


@login_required
def form16_main(request):
    """Главная страница Формы 16"""
    articles_count = Form16Article.objects.filter(
        user=request.user, is_active=True
    ).count()

    context = {
        "articles_count": articles_count,
        "has_articles": articles_count > 0,
        "max_positions": 50,
    }
    return render(request, "forms_app/form16_main.html", context)


@login_required
def form16_edit_table(request):
    """Редактирование таблицы артикулов"""
    # Получаем существующие артикулы пользователя
    existing_articles = {
        article.position: article
        for article in Form16Article.objects.filter(user=request.user)
    }

    if request.method == "POST":
        try:
            with transaction.atomic():
                # Обрабатываем все 50 позиций
                for position in range(1, 51):
                    article_wb = request.POST.get(f"article_wb_{position}", "").strip()
                    our_article = request.POST.get(
                        f"our_article_{position}", ""
                    ).strip()
                    comments = request.POST.get(f"comments_{position}", "").strip()
                    is_active = request.POST.get(f"active_{position}") == "on"

                    if position in existing_articles:
                        # Обновляем существующий артикул
                        article = existing_articles[position]
                        if article_wb:
                            article.article_wb = article_wb
                            article.our_article = our_article
                            article.comments = comments
                            article.is_active = is_active
                            article.save()
                        else:
                            # Если поле пустое - удаляем артикул
                            article.delete()
                    else:
                        # Создаем новый артикул
                        if article_wb:
                            Form16Article.objects.create(
                                user=request.user,
                                position=position,
                                article_wb=article_wb,
                                our_article=our_article,
                                comments=comments,
                                is_active=is_active,
                            )

            messages.success(request, "Таблица артикулов успешно сохранена!")
            return redirect("forms_app:form16_edit_table")

        except Exception as e:
            messages.error(request, f"Ошибка при сохранении: {str(e)}")

    # Для GET запроса - готовим данные для формы
    articles_list = Form16Article.objects.filter(user=request.user).order_by("position")

    # Создаем список данных для каждой позиции
    table_data = []
    for position in range(1, 51):
        if position in existing_articles:
            article = existing_articles[position]
            table_data.append(
                {
                    "position": position,
                    "article_wb": article.article_wb,
                    "our_article": article.our_article,
                    "comments": article.comments,
                    "is_active": article.is_active,
                    "exists": True,
                }
            )
        else:
            table_data.append(
                {
                    "position": position,
                    "article_wb": "",
                    "our_article": "",
                    "comments": "",
                    "is_active": False,
                    "exists": False,
                }
            )

    context = {"table_data": table_data, "articles_list": articles_list}
    return render(request, "forms_app/form16_edit_table.html", context)


@login_required
def form16_generate_report(request):
    """Генерация отчета на основе загруженного файла и сохраненных артикулов"""
    # Получаем активные артикулы пользователя
    articles_qs = Form16Article.objects.filter(
        user=request.user, is_active=True
    ).order_by("position")

    # Получаем артикулы как строки и очищаем от лишних пробелов
    articles = [str(article.article_wb).strip() for article in articles_qs]

    if not articles:
        messages.error(
            request, "У вас нет сохраненных артикулов. Заполните таблицу сначала."
        )
        return redirect("forms_app:form16_edit_table")

    # Инициализируем форму
    form = None

    if request.method == "POST":
        # Проверяем наличие файла
        if "file" not in request.FILES:
            messages.error(request, "Файл не загружен")
            return redirect("forms_app:form16_generate_report")

        file = request.FILES["file"]

        # Проверяем расширение файла
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Файл должен быть в формате .xlsx")
            return redirect("forms_app:form16_generate_report")

        try:
            # Сначала проверим, какие страницы есть в файле
            xls = pd.ExcelFile(file)
            sheet_names = xls.sheet_names

            # Ищем нужную страницу
            target_sheet = None
            for sheet in sheet_names:
                if "детальная" in sheet.lower():
                    target_sheet = sheet
                    break

            if not target_sheet:
                target_sheet = sheet_names[0]
                messages.warning(
                    request,
                    f"Страница 'Детальная информация' не найдена. Используется первая страница: '{target_sheet}'",
                )

            # Читаем файл
            df = pd.read_excel(
                file,
                sheet_name=target_sheet,
                header=1,
            )

            # Преобразуем колонку 'Артикул WB' в строки
            df["Артикул WB_clean"] = df["Артикул WB"].astype(str).str.strip()

            # Очищаем наши артикулы
            articles_clean = [str(article).strip() for article in articles]

            # === Поиск артикулов в файле по ПОЛНОМУ совпадению Артикул WB ===
            # Получаем все уникальные артикулы из файла
            all_articles_in_file = df["Артикул WB_clean"].unique().tolist()

            # Проверяем каждый полный артикул
            found_articles = []
            not_found_articles = []

            for article in articles_clean:
                if article in all_articles_in_file:
                    found_articles.append(article)
                else:
                    not_found_articles.append(article)

            found_count = len(found_articles)

            if found_count == 0:
                error_msg = f"❌ Ни один артикул не найден в файле!\n\n"
                error_msg += f"Искали артикулы: {articles_clean}\n\n"
                error_msg += f"Артикулы в файле (первые 30):\n"
                for i, art in enumerate(all_articles_in_file[:30], 1):
                    error_msg += f"{i}. {art}\n"

                messages.error(request, error_msg)
                return redirect("forms_app:form16_generate_report")

            # Фильтруем найденные артикулы по полному совпадению
            df_filtered = df[df["Артикул WB_clean"].isin(articles_clean)].copy()

            # Создаем порядок сортировки (по нашему списку артикулов)
            article_order = {article: i for i, article in enumerate(articles_clean)}

            # Добавляем колонку для сортировки
            df_filtered["Порядок_артикула"] = df_filtered["Артикул WB_clean"].map(
                article_order
            )
            df_filtered["Порядок_артикула"] = df_filtered["Порядок_артикула"].fillna(
                999
            )

            # Преобразуем размер для сортировки
            def try_convert_to_numeric(x):
                try:
                    return float(x)
                except:
                    return float("inf")

            df_filtered["Размер_число"] = (
                df_filtered["Размер"].astype(str).apply(try_convert_to_numeric)
            )

            # Сортируем
            df_sorted = df_filtered.sort_values(
                ["Порядок_артикула", "Артикул продавца", "Размер_число"],
                ascending=[True, True, True],
            )

            # Выбираем нужные колонки
            available_columns = []
            desired_columns = [
                "Артикул продавца",
                "Артикул WB",
                "Размер",
                "Доступность",
                "Комментарии",
                "Заказали, шт",
                "Выкупили, шт",
                "Процент выкупа",
                "Остатки на текущий день, шт",
            ]

            for col in desired_columns:
                if col in df_sorted.columns:
                    available_columns.append(col)

            df_result = df_sorted[available_columns].copy()

            # Удаляем дубликаты
            df_result = df_result.drop_duplicates()

            # === Комментарии по ПОЛНОМУ Артикул WB ===
            # Создаем словарь комментариев по артикулам WB из нашей базы
            comments_dict = {}
            for article in Form16Article.objects.filter(
                user=request.user, is_active=True
            ):
                if article.comments:
                    comments_dict[str(article.article_wb).strip()] = article.comments

            # Функция для получения комментариев
            def get_comments(row):
                article_wb = (
                    str(row["Артикул WB"]).strip() if "Артикул WB" in row else ""
                )
                return comments_dict.get(article_wb, "")

            # === Остатки по первым 4 символам Артикул продавца ===
            # Получаем все записи остатков пользователя из StockRecord
            from forms_app.models import StockRecord

            stock_records = StockRecord.objects.filter(user=request.user)

            # Создаем словари для быстрого поиска остатков по первым 4 символам
            stock_by_prefix_and_size = {}
            stock_by_prefix_only = {}

            for record in stock_records:
                # Полный артикул продавца из нашей базы (article_full_name)
                full_article = (
                    str(record.article_full_name).strip()
                    if record.article_full_name
                    else ""
                )
                size_key = str(record.size).strip() if record.size else ""

                # Извлекаем префикс (первые 4 символа) из артикула продавца
                article_prefix = (
                    full_article[:4] if len(full_article) >= 4 else full_article
                )

                # Составляем ключ для точного совпадения по префиксу и размеру
                if article_prefix and size_key:
                    composite_key = f"{article_prefix}_{size_key}"
                    if composite_key not in stock_by_prefix_and_size:
                        stock_by_prefix_and_size[composite_key] = 0
                    stock_by_prefix_and_size[composite_key] += record.quantity or 0

                # Также сохраняем сумму по префиксу (на случай если размер не совпадает)
                if article_prefix:
                    if article_prefix not in stock_by_prefix_only:
                        stock_by_prefix_only[article_prefix] = 0
                    stock_by_prefix_only[article_prefix] += record.quantity or 0

            # Функция для получения остатков по префиксу (первые 4 символа Артикул продавца)
            def get_our_stock_by_prefix(row):
                # Получаем артикул продавца из строки (из файла WB)
                article_vendor = (
                    str(row.get("Артикул продавца", "")).strip()
                    if "Артикул продавца" in row
                    else ""
                )
                size = str(row.get("Размер", "")).strip() if "Размер" in row else ""

                if not article_vendor:
                    return 0

                # Извлекаем префикс (первые 4 символа) из артикула продавца из файла WB
                article_prefix = (
                    article_vendor[:4] if len(article_vendor) >= 4 else article_vendor
                )

                # Сначала ищем точное совпадение по префиксу и размеру
                composite_key = f"{article_prefix}_{size}"
                if composite_key in stock_by_prefix_and_size:
                    return stock_by_prefix_and_size[composite_key]

                # Если точного совпадения нет, ищем по префиксу без учета размера
                if article_prefix in stock_by_prefix_only:
                    return stock_by_prefix_only[article_prefix]

                # Если ничего не найдено
                return 0

            # Добавляем колонку "Комментарии" после "Доступность"
            if "Доступность" in df_result.columns:
                availability_idx = df_result.columns.get_loc("Доступность")
                df_result.insert(availability_idx + 1, "Комментарии", "")
            else:
                df_result["Комментарии"] = ""

            # Заполняем комментарии
            df_result["Комментарии"] = df_result.apply(get_comments, axis=1)

            # Добавляем пустую колонку "Отгрузка ФБО"
            df_result["Отгрузка ФБО"] = ""

            # Добавляем колонку "Наши остатки" после "Отгрузка ФБО"
            if "Отгрузка ФБО" in df_result.columns:
                fbo_idx = df_result.columns.get_loc("Отгрузка ФБО")
                df_result.insert(fbo_idx + 1, "Наши остатки", 0)
            else:
                df_result["Наши остатки"] = 0

            # Заполняем значения остатков
            df_result["Наши остатки"] = df_result.apply(get_our_stock_by_prefix, axis=1)

            # Переупорядочиваем колонки
            columns_order = [
                "Артикул продавца",
                "Артикул WB",
                "Размер",
                "Доступность",
                "Комментарии",
                "Заказали, шт",
                "Выкупили, шт",
                "Процент выкупа",
                "Остатки на текущий день, шт",
                "Отгрузка ФБО",
                "Наши остатки",
            ]

            # Оставляем только существующие колонки в нужном порядке
            final_columns = [col for col in columns_order if col in df_result.columns]
            df_result = df_result[final_columns]

            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"ФОРМИРОВАНИЕ_ФБО"

            # === ЗАГОЛОВКИ ТАБЛИЦЫ ===
            headers = list(df_result.columns)
            ws.append(headers)
            HEADER_ROW = ws.max_row  # Запоминаем строку с заголовками

            # === ДАННЫЕ ТАБЛИЦЫ ===
            for _, row in df_result.iterrows():
                ws.append(row.tolist())

            # Начало данных (первая строка после заголовков)
            DATA_START_ROW = HEADER_ROW + 1

            # === ФОРМАТИРОВАНИЕ ===

            # 1. Заголовки таблицы
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=HEADER_ROW, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # 2. Данные таблицы (цветовое кодирование по артикулам)
            colors = [
                PatternFill(
                    start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"
                ),
                PatternFill(
                    start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"
                ),
                PatternFill(
                    start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"
                ),
                PatternFill(
                    start_color="F8CECC", end_color="F8CECC", fill_type="solid"
                ),
                PatternFill(
                    start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"
                ),
                PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                ),
                PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                ),
                PatternFill(
                    start_color="D0E0E3", end_color="D0E0E3", fill_type="solid"
                ),
                PatternFill(
                    start_color="E8D5E7", end_color="E8D5E7", fill_type="solid"
                ),
                PatternFill(
                    start_color="D4E6F1", end_color="D4E6F1", fill_type="solid"
                ),
                PatternFill(
                    start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
                ),
                PatternFill(
                    start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
                ),
                PatternFill(
                    start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"
                ),
                PatternFill(
                    start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"
                ),
                PatternFill(
                    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                ),
            ]

            # Раскрашиваем строки данных по артикулам
            current_article = None
            color_index = 0
            article_colors = {}

            for row in range(DATA_START_ROW, ws.max_row + 1):
                article = ws.cell(
                    row=row, column=2
                ).value  # Колонка 'Артикул WB' (вторая колонка)
                if article:
                    article_str = str(article).strip()
                    if article_str != current_article:
                        current_article = article_str
                        if current_article not in article_colors:
                            article_colors[current_article] = colors[
                                color_index % len(colors)
                            ]
                            color_index += 1

                    fill_color = article_colors.get(current_article, colors[0])

                    # Форматируем всю строку данных
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill_color
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                        # Выравнивание для числовых колонок
                        if col in [
                            6,
                            7,
                            8,
                            9,
                            11,
                        ]:  # Индексы числовых колонок
                            try:
                                if cell.value is not None and str(cell.value).strip():
                                    cell.alignment = Alignment(
                                        horizontal="right", vertical="center"
                                    )
                            except:
                                cell.alignment = Alignment(
                                    horizontal="left", vertical="center"
                                )
                        else:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )

            # 3. Устанавливаем ширину для всех колонок
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                # Проверяем заголовок
                header_cell = ws.cell(row=HEADER_ROW, column=col_idx)
                max_length = max(max_length, len(str(header_cell.value or "")))

                # Проверяем данные
                for row in range(DATA_START_ROW, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        cell_value = str(cell.value or "")
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass

                # Устанавливаем ширину
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 4. Особенная ширина для колонок "Отгрузка ФБО" и "Наши остатки"
            if len(headers) >= 10:
                last_col_letter = get_column_letter(len(headers))
                prev_col_letter = get_column_letter(len(headers) - 1)
                ws.column_dimensions[prev_col_letter].width = 18  # Отгрузка ФБО
                ws.column_dimensions[last_col_letter].width = 15  # Наши остатки

            # Сохраняем в буфер
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Возвращаем файл
            filename = f"ОБОРАЧИВАЕМОСТЬ_Форма16_{found_count}_артикулов.xlsx"
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=filename,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            return response

        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")
            import traceback

            traceback.print_exc()

    # Для GET запроса
    from forms_app.forms import Form16UploadForm

    form = Form16UploadForm()

    context = {
        "form": form,
        "articles": articles_qs,
        "articles_count": len(articles),
    }

    return render(request, "forms_app/form16_generate_report.html", context)


@login_required
def form16_delete_all(request):
    """Удаление всех артикулов пользователя"""
    if request.method == "POST":
        count = Form16Article.objects.filter(user=request.user).count()
        Form16Article.objects.filter(user=request.user).delete()
        messages.success(request, f"Удалено {count} артикулов")
        return redirect("forms_app:form16_edit_table")

    return render(
        request,
        "forms_app/form16_delete_all.html",
        {"articles_count": Form16Article.objects.filter(user=request.user).count()},
    )
